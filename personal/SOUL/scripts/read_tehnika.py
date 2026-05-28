#!/usr/bin/env python3
"""Читает технику из двух источников:
1) Лист `Техника` в смете .xlsx (Drive 1CrD046MP...)
2) Лист `ТЕХНИКА + МЕБЕЛЬ` в ТЗ Google Sheet (19xa4EdE...)
Печатает оба массива построчно с непустыми ячейками.
"""
import warnings, io, json
warnings.filterwarnings("ignore", category=FutureWarning)

from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

KEY = "/Users/pro2kuror/.config/vibecoding/assistant/secrets/finance-director-sheets-903611b799c3.json"
SMETA_FILE_ID = "1CrD046MP0Rxg-n0aBaNUW3tWs_cS3aOB"
TZ_SHEET_ID = "19xa4EdEN6ih9AoHYzMGg9D4ALLgxetOu8qZsfsbu5w8"
TZ_SHEET_NAME = "ТЕХНИКА + МЕБЕЛЬ"

creds = Credentials.from_service_account_file(
    KEY, scopes=[
        "https://www.googleapis.com/auth/spreadsheets.readonly",
        "https://www.googleapis.com/auth/drive.readonly",
    ]
)
drive = build("drive", "v3", credentials=creds, cache_discovery=False)
sheets = build("sheets", "v4", credentials=creds, cache_discovery=False)

# === 1. Лист `Техника` из xlsx ===
print("=" * 80)
print("ИСТОЧНИК 1: Лист `Техника` в смете .xlsx")
print("=" * 80)

req = drive.files().get_media(fileId=SMETA_FILE_ID)
buf = io.BytesIO()
dl = MediaIoBaseDownload(buf, req)
done = False
while not done:
    _, done = dl.next_chunk()
buf.seek(0)
out_path = "/Users/pro2kuror/Desktop/VibeCoding/personal/SOUL/cache/smeta.xlsx"
with open(out_path, "wb") as f:
    f.write(buf.read())

import openpyxl
wb = openpyxl.load_workbook(out_path, data_only=True)
if "Техника" in wb.sheetnames:
    ws = wb["Техника"]
    rows = []
    for r in ws.iter_rows(values_only=True):
        nonempty = [str(c) if c is not None else "" for c in r]
        if any(s.strip() for s in nonempty):
            rows.append(nonempty)
    print(f"Всего строк с данными: {len(rows)}\n")
    for i, row in enumerate(rows, 1):
        print(f"  {i:2d} | " + " | ".join(s for s in row[:8] if s.strip()))
else:
    print("Лист `Техника` не найден")

# === 2. Лист `ТЕХНИКА + МЕБЕЛЬ` из ТЗ ===
print()
print("=" * 80)
print(f"ИСТОЧНИК 2: Лист `{TZ_SHEET_NAME}` в ТЗ Google Sheet")
print("=" * 80)

vals = sheets.spreadsheets().values().get(
    spreadsheetId=TZ_SHEET_ID, range=f"{TZ_SHEET_NAME}!A1:N100",
    valueRenderOption="FORMATTED_VALUE"
).execute().get("values", [])
print(f"Всего строк: {len(vals)}\n")
for i, row in enumerate(vals, 1):
    nonempty = [str(c) for c in row if c is not None and str(c).strip()]
    if nonempty:
        print(f"  {i:2d} | " + " | ".join(nonempty[:8]))
