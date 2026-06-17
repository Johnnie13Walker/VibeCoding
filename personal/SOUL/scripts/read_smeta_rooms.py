#!/usr/bin/env python3
"""Прочитать листы сметы по жилым комнатам, найти сантехнику."""
import warnings, json, io, re
warnings.filterwarnings("ignore", category=FutureWarning)

from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

KEY = "/Users/pro2kuror/.config/vibecoding/assistant/secrets/finance-director-sheets-903611b799c3.json"
SMETA_FILE_ID = "1CrD046MP0Rxg-n0aBaNUW3tWs_cS3aOB"
TZ_SHEET_ID = "19xa4EdEN6ih9AoHYzMGg9D4ALLgxetOu8qZsfsbu5w8"

creds = Credentials.from_service_account_file(
    KEY, scopes=[
        "https://www.googleapis.com/auth/spreadsheets.readonly",
        "https://www.googleapis.com/auth/drive.readonly",
    ]
)
drive = build("drive", "v3", credentials=creds, cache_discovery=False)
sheets = build("sheets", "v4", credentials=creds, cache_discovery=False)

# Скачиваем .xlsx
print("=== Загрузка сметы .xlsx ===")
req = drive.files().get_media(fileId=SMETA_FILE_ID)
buf = io.BytesIO()
downloader = MediaIoBaseDownload(buf, req)
done = False
while not done:
    _, done = downloader.next_chunk()
buf.seek(0)
with open("/tmp/smeta.xlsx","wb") as f:
    f.write(buf.read())
print("OK, /tmp/smeta.xlsx")

import openpyxl
wb = openpyxl.load_workbook("/tmp/smeta.xlsx", data_only=True)
print(f"\nЛисты сметы: {wb.sheetnames}")

# Соответствие комнат (поищем что подходит)
target_rooms = ["кухня", "гостиная", "спальня", "детская", "кабинет"]
matched_sheets = []
for s in wb.sheetnames:
    sl = s.lower()
    for t in target_rooms:
        if t in sl:
            matched_sheets.append(s); break
print(f"\nЦелевые листы: {matched_sheets}")

# Ключевые слова сантехники
santech_kw = ["смесител", "кран", "сифон", "трап", "лоток", "слив", "раковин", "мойк",
              "душев", "ванн", "унитаз", "биде", "полотенцесуш", "счётчик воды",
              "смыв", "инсталл", "коллектор", "редуктор", "бачок", "мыльниц"]

print("\n" + "=" * 80)
for sname in matched_sheets:
    print(f"\n##### Лист: {sname} #####")
    ws = wb[sname]
    rows_with_santech = []
    for r in ws.iter_rows(values_only=True):
        line = " ".join(str(c) for c in r if c is not None).lower()
        if not line.strip(): continue
        if any(kw in line for kw in santech_kw):
            # сохранить непустые ячейки
            nonempty = [str(c) for c in r if c is not None and str(c).strip()]
            rows_with_santech.append(nonempty)
    if rows_with_santech:
        print(f"  Найдено {len(rows_with_santech)} строк со словами сантехники:")
        for row in rows_with_santech:
            print("   |", " | ".join(row[:8]))
    else:
        print("  (сантехника не найдена)")
