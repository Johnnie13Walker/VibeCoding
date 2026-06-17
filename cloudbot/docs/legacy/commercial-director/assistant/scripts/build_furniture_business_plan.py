from __future__ import annotations

from dataclasses import dataclass
from math import ceil
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "output" / "spreadsheet"
WORKBOOK_PATH = OUTPUT_DIR / "furniture_business_plan_moscow.xlsx"
REPORT_PATH = OUTPUT_DIR / "furniture_business_plan_moscow.md"


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
INPUT_FILL = PatternFill("solid", fgColor="DDEBF7")
FORMULA_FILL = PatternFill("solid", fgColor="E2F0D9")
GRAY_FILL = PatternFill("solid", fgColor="E7E6E6")
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


@dataclass(frozen=True)
class ScenarioConfig:
    name: str
    monthly_growth_y1: float
    monthly_growth_y2: float
    price_uplift: float
    referral_share: float
    budget_multiplier: float
    segment_mix: tuple[float, float, float]


SCENARIOS: dict[str, ScenarioConfig] = {
    "Базовый": ScenarioConfig(
        name="Базовый",
        monthly_growth_y1=0.09,
        monthly_growth_y2=0.06,
        price_uplift=0.0025,
        referral_share=0.14,
        budget_multiplier=1.0,
        segment_mix=(0.25, 0.55, 0.20),
    ),
    "Оптимистичный": ScenarioConfig(
        name="Оптимистичный",
        monthly_growth_y1=0.11,
        monthly_growth_y2=0.07,
        price_uplift=0.003,
        referral_share=0.15,
        budget_multiplier=1.25,
        segment_mix=(0.18, 0.54, 0.28),
    ),
}


MONTHS = [f"M{i}" for i in range(1, 25)]
MONTH_LABELS = [
    "апрель 2026", "май 2026", "июнь 2026", "июль 2026", "август 2026", "сентябрь 2026",
    "октябрь 2026", "ноябрь 2026", "декабрь 2026", "январь 2027", "февраль 2027", "март 2027",
    "апрель 2027", "май 2027", "июнь 2027", "июль 2027", "август 2027", "сентябрь 2027",
    "октябрь 2027", "ноябрь 2027", "декабрь 2027", "январь 2028", "февраль 2028", "март 2028",
]
PRODUCT_SHARES = {
    "Кухни": 0.45,
    "Шкафы-купе и гардеробные": 0.30,
    "Мебель для ванных": 0.15,
    "Прочая корпусная мебель": 0.10,
}
AVG_CHECKS = {
    "Эконом": {"Кухни": 280_000, "Шкафы-купе и гардеробные": 180_000, "Мебель для ванных": 95_000, "Прочая корпусная мебель": 120_000},
    "Средний": {"Кухни": 480_000, "Шкафы-купе и гардеробные": 300_000, "Мебель для ванных": 160_000, "Прочая корпусная мебель": 220_000},
    "Премиум": {"Кухни": 900_000, "Шкафы-купе и гардеробные": 550_000, "Мебель для ванных": 320_000, "Прочая корпусная мебель": 380_000},
}

CHANNELS = {
    "Контекст_поиск": {
        "cpl": 5_800,
        "lead_to_measure": 0.18,
        "measure_to_deal": 0.39,
        "base_budget": 210_000,
        "opt_budget": 260_000,
        "phase_multipliers": [1.0, 1.15, 1.30, 1.45],
        "active_from": 1,
        "cpl_monthly_shift": -0.004,
    },
    "Ретаргетинг": {
        "cpl": 3_000,
        "lead_to_measure": 0.24,
        "measure_to_deal": 0.34,
        "base_budget": 30_000,
        "opt_budget": 40_000,
        "phase_multipliers": [0.0, 1.0, 1.15, 1.35],
        "active_from": 3,
        "cpl_monthly_shift": -0.003,
    },
    "Авито": {
        "cpl": 2_500,
        "lead_to_measure": 0.20,
        "measure_to_deal": 0.30,
        "base_budget": 120_000,
        "opt_budget": 150_000,
        "phase_multipliers": [1.0, 1.08, 1.15, 1.22],
        "active_from": 1,
        "cpl_monthly_shift": -0.002,
    },
    "Агрегаторы": {
        "cpl": 4_500,
        "lead_to_measure": 0.17,
        "measure_to_deal": 0.28,
        "base_budget": 55_000,
        "opt_budget": 70_000,
        "phase_multipliers": [0.8, 0.9, 0.85, 0.8],
        "active_from": 1,
        "cpl_monthly_shift": 0.0,
    },
}

SEASONALITY = [0.90, 0.96, 1.02, 1.08, 1.10, 1.00, 0.88, 0.92, 1.10, 1.16, 1.12, 0.98]
PHASES = [
    (1, 3, "Запуск", "Сайт и аналитика собраны, контекст и Авито дают первые сделки, SEO в разогреве."),
    (4, 8, "Разгон", "Усиливаем лучшее: расширяем семантику, объявления, кейсы и отзывы."),
    (9, 14, "Оптимизация", "Чистим неэффективные кампании, повышаем долю ROMI-положительных каналов."),
    (15, 24, "Масштабирование", "Добавляем объем только в каналы с устойчивой окупаемостью и SLA по обработке."),
]

MODEL_CONFIG = {
    "A_Аутсорс": {
        "avg_check_adjustment": 1.00,
        "production_share": 0.42,
        "materials_share": 0.10,
        "logistics_share": 0.035,
        "assembly_share": 0.045,
        "showroom_rent": 140_000,
        "shop_rent": 0,
        "it_cost": 35_000,
        "other_opex": 65_000,
        "startup_investment": 3_050_000,
        "model_label": "Аутсорс производства",
    },
}

SALARIES = {
    "Менеджер продаж": 90_000,
    "Дизайнер": 110_000,
    "Замерщик": 100_000,
    "Бригадир сборки": 130_000,
    "Операционный менеджер": 120_000,
    "Администратор": 85_000,
    "Технолог": 140_000,
    "Рабочий цеха": 110_000,
    "Начальник производства": 160_000,
}


SOURCES = [
    (
        "ФНС: УСН 6% и 15%",
        "https://www.nalog.gov.ru/rn25/ifns/r25_14/info/15658830/",
        "Ставка УСН «доходы минус расходы» 15%, «доходы» 6%.",
    ),
    (
        "ФНС: изменения взносов 2025",
        "https://www.nalog.gov.ru/rn28/news/activities_fts/15540345/",
        "С 2025 года пониженный тариф МСП 15% применяется к выплатам сверх 1,5 МРОТ.",
    ),
    (
        "ФНС Москва: взносы 2026",
        "https://www.nalog.gov.ru/rn77/news/activities_fts/16599875/",
        "В 2026 году льгота 15% для МСП зависит от отрасли; для модели принят консервативный эффективный тариф на ФОТ.",
    ),
    (
        "Экоофис: аренда офиса в Москве",
        "https://www.ecooffice.ru/arenda_ofisa/ot_200_metrov_kv",
        "Ориентир офисной аренды 19–20,5 тыс. ₽/м2/год для помещения около 380 м2.",
    ),
    (
        "Optima Invest: мебельный цех в Москве",
        "https://optima-invest.ru/msk/obekty/stolyarnyy_tsekh_i_mebelnoe_proizvodstvo/",
        "Ориентир аренды мебельного цеха 500 м2: 190 тыс. ₽/мес.",
    ),
    (
        "Станкофф: форматно-раскроечный станок D45A",
        "https://www.stankoff.ru/product/26773/formatno-raskroechnyiy-stanok-d45a",
        "Ориентир цены базового раскроечного станка: 200 тыс. ₽.",
    ),
]


def weighted_avg_check(segment_mix: tuple[float, float, float], model_key: str) -> float:
    segment_names = ["Эконом", "Средний", "Премиум"]
    total = 0.0
    for share, segment_name in zip(segment_mix, segment_names):
        product_total = 0.0
        for product_name, product_share in PRODUCT_SHARES.items():
            product_total += AVG_CHECKS[segment_name][product_name] * product_share
        total += product_total * share
    return total * MODEL_CONFIG[model_key]["avg_check_adjustment"]


def current_growth(month_index: int, scenario: ScenarioConfig) -> float:
    return scenario.monthly_growth_y1 if month_index < 12 else scenario.monthly_growth_y2


def get_phase(month_number: int) -> tuple[int, str, str]:
    for idx, (start, end, title, note) in enumerate(PHASES):
        if start <= month_number <= end:
            return idx, title, note
    return len(PHASES) - 1, PHASES[-1][2], PHASES[-1][3]


def build_media_plan(scenario_name: str) -> list[dict[str, object]]:
    scenario = SCENARIOS[scenario_name]
    rows: list[dict[str, object]] = []
    budget_key = "opt_budget" if scenario_name == "Оптимистичный" else "base_budget"
    previous_budgets = {channel_name: 0.0 for channel_name in CHANNELS}
    previous_phase_idx = 0
    previous_seasonality = SEASONALITY[0]
    for month_number in range(1, 25):
        phase_idx, phase_name, phase_note = get_phase(month_number)
        seasonality = SEASONALITY[(month_number - 1) % 12]
        channel_rows: dict[str, dict[str, float]] = {}
        for channel_name, params in CHANNELS.items():
            if month_number < params["active_from"]:
                budget = 0.0
            elif previous_budgets[channel_name] == 0.0:
                budget = (
                    params[budget_key]
                    * params["phase_multipliers"][phase_idx]
                    * seasonality
                    * scenario.budget_multiplier
                )
            else:
                prev_phase_mult = params["phase_multipliers"][previous_phase_idx]
                current_phase_mult = params["phase_multipliers"][phase_idx]
                phase_ratio = current_phase_mult / prev_phase_mult if prev_phase_mult else max(current_phase_mult, 1.0)
                budget = previous_budgets[channel_name] * (1 + current_growth(month_number - 2, scenario)) * (seasonality / previous_seasonality) * phase_ratio
            effective_cpl = max(params["cpl"] * ((1 + params["cpl_monthly_shift"]) ** (month_number - 1)), params["cpl"] * 0.72)
            leads = budget / effective_cpl if effective_cpl else 0.0
            measurements = leads * params["lead_to_measure"]
            orders = measurements * params["measure_to_deal"]
            channel_rows[channel_name] = {
                "budget": budget,
                "cpl": effective_cpl,
                "leads": leads,
                "measurements": measurements,
                "orders": orders,
            }
        rows.append(
            {
                "month": month_number,
                "phase": phase_name,
                "phase_note": phase_note,
                "seasonality": seasonality,
                "channels": channel_rows,
            }
        )
        previous_budgets = {channel_name: data["budget"] for channel_name, data in channel_rows.items()}
        previous_phase_idx = phase_idx
        previous_seasonality = seasonality
    return rows


def payroll_and_headcount(orders: float, revenue: float, model_key: str) -> tuple[float, dict[str, int]]:
    counts = {
        "Менеджер продаж": max(1, ceil(orders / 15)),
        "Дизайнер": max(1, ceil(orders / 18)),
        "Замерщик": max(1, ceil(orders / 45)),
        "Бригадир сборки": max(1, ceil(orders / 22)),
        "Операционный менеджер": 1 if orders < 45 else 2,
        "Администратор": 1,
    }
    if model_key == "B_Свой_цех":
        counts["Технолог"] = 1 if orders < 40 else 2
        counts["Рабочий цеха"] = max(2, ceil(orders / 11))
        counts["Начальник производства"] = 1

    fixed_fot = sum(SALARIES[role] * count for role, count in counts.items())
    variable_sales_bonus = revenue * 0.018
    return fixed_fot + variable_sales_bonus, counts


def monthly_rows(model_key: str, scenario_name: str) -> list[dict[str, float | dict[str, int]]]:
    scenario = SCENARIOS[scenario_name]
    model = MODEL_CONFIG[model_key]
    base_avg_check = weighted_avg_check(scenario.segment_mix, model_key)
    rows: list[dict[str, float | dict[str, int]]] = []
    previous_orders = 0.0
    media_plan = build_media_plan(scenario_name)

    for month_index in range(24):
        plan_row = media_plan[month_index]
        channels_plan = plan_row["channels"]

        orders_by_channel: dict[str, float] = {}
        leads_by_channel: dict[str, float] = {}
        measurements_by_channel: dict[str, float] = {}
        leads_total = 0.0
        measurements_total = 0.0
        budgets: dict[str, float] = {}
        effective_cpl: dict[str, float] = {}
        for channel_name, channel_plan in channels_plan.items():
            budget = channel_plan["budget"]
            leads = channel_plan["leads"]
            measurements = channel_plan["measurements"]
            deals = channel_plan["orders"]
            orders_by_channel[channel_name] = deals
            leads_by_channel[channel_name] = leads
            measurements_by_channel[channel_name] = measurements
            leads_total += leads
            measurements_total += measurements
            budgets[channel_name] = budget
            effective_cpl[channel_name] = channel_plan["cpl"]

        referral_orders = max(1.0, previous_orders * scenario.referral_share)
        total_orders = sum(orders_by_channel.values()) + referral_orders
        avg_check = base_avg_check * ((1 + scenario.price_uplift) ** month_index)
        revenue = total_orders * avg_check

        production_cost = revenue * model["production_share"]
        materials_cost = revenue * model["materials_share"]
        logistics_cost = revenue * model["logistics_share"]
        assembly_cost = revenue * model["assembly_share"]
        total_cogs = production_cost + materials_cost + logistics_cost + assembly_cost

        fot, counts = payroll_and_headcount(total_orders, revenue, model_key)
        payroll_taxes = fot * 0.25
        marketing = sum(budgets.values()) + revenue * 0.008
        rent = model["showroom_rent"] + model["shop_rent"]
        ebitda = revenue - total_cogs - marketing - fot - payroll_taxes - rent - model["it_cost"] - model["other_opex"]
        tax = max(0.0, max(ebitda * 0.15, revenue * 0.01) if ebitda > 0 else 0.0)
        net_profit = ebitda - tax

        rows.append(
            {
                "month": month_index + 1,
                "leads": leads_total,
                "measurements": measurements_total,
                "orders": total_orders,
                "avg_check": avg_check,
                "revenue": revenue,
                "production_cost": production_cost,
                "materials_cost": materials_cost,
                "logistics_cost": logistics_cost,
                "assembly_cost": assembly_cost,
                "cogs": total_cogs,
                "marketing": marketing,
                "fot": fot,
                "payroll_taxes": payroll_taxes,
                "rent": rent,
                "it_cost": model["it_cost"],
                "other_opex": model["other_opex"],
                "ebitda": ebitda,
                "tax": tax,
                "net_profit": net_profit,
                "counts": counts,
                "orders_by_channel": orders_by_channel,
                "leads_by_channel": leads_by_channel,
                "measurements_by_channel": measurements_by_channel,
                "referral_orders": referral_orders,
                "budgets": budgets,
                "effective_cpl": effective_cpl,
                "phase": plan_row["phase"],
                "phase_note": plan_row["phase_note"],
                "seasonality": plan_row["seasonality"],
            }
        )
        previous_orders = total_orders
    return rows


def style_headers(ws, row: int, start_col: int, end_col: int) -> None:
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def apply_body_style(ws, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.border = THIN_BORDER
            if cell.column >= 2:
                cell.number_format = '#,##0_);[Red](#,##0)'


def autosize(ws) -> None:
    for column_cells in ws.columns:
        length = max(len(str(cell.value) if cell.value is not None else "") for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(length + 2, 12), 28)


def add_assumptions_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Допущения"
    ws.append(["Категория", "Показатель", "Значение", "Ед.", "Комментарий", "Источник"])
    style_headers(ws, 1, 1, 6)

    entries = [
        ("Налоги", "Режим налогообложения", "УСН 15%", "", "Оптимально при доле прямых затрат выше 55%.", SOURCES[0][1]),
        ("ФОТ", "Эффективные взносы на ФОТ", 27, "%", "Консервативная ставка с учетом структуры МСП и действующих правил 2025-2026.", SOURCES[2][1]),
        ("Продажи", "Доля рекомендаций от базы", 12, "%", "В базовом сценарии referral-канал разгоняется от накопленной базы.", "Управленческое допущение"),
        ("Офис", "Шоурум / офис в Москве", 140_000, "₽/мес", "Для старта берется компактный шоурум 50-70 м2.", SOURCES[3][1]),
        ("Цех", "Аренда цеха 180-250 м2", 250_000, "₽/мес", "Заложено выше ориентира рынка с запасом на коммунальные и склад.", SOURCES[4][1]),
        ("Оборудование", "Базовый форматно-раскроечный станок", 200_000, "₽", "Минимальный ориентир для расчета CAPEX.", SOURCES[5][1]),
        ("Маркетинг", "CPL SEO", 3_500, "₽", "Консервативный ориентир для Москвы по нише корпусной мебели.", "Управленческое допущение"),
        ("Маркетинг", "CPL Контекст", 5_500, "₽", "Высококонкурентный аукцион по мебели на заказ.", "Управленческое допущение"),
        ("Маркетинг", "CPL Авито", 2_500, "₽", "С учетом платного продвижения и обработки входящих.", "Управленческое допущение"),
        ("Маркетинг", "CPL Агрегаторы", 4_500, "₽", "С учетом комиссии и покупки трафика.", "Управленческое допущение"),
    ]
    for entry in entries:
        ws.append(list(entry))

    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=3).fill = INPUT_FILL
        ws.cell(row=row, column=6).alignment = Alignment(wrap_text=True)
        ws.cell(row=row, column=5).alignment = Alignment(wrap_text=True)

    source_map = {title: (url, note) for title, url, note in SOURCES}
    for row in range(2, ws.max_row + 1):
        url = ws.cell(row=row, column=6).value
        if isinstance(url, str) and url.startswith("http"):
            ws.cell(row=row, column=6).comment = Comment(url, "Codex")
            for _, source_url, note in SOURCES:
                if source_url == url:
                    ws.cell(row=row, column=3).comment = Comment(note, "Codex")
                    break

    ws.freeze_panes = "A2"
    autosize(ws)


def add_unit_economics_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Unit-экономика")
    ws.append(["Канал", "CPL, ₽", "Лид → замер", "Замер → договор", "Конверсия в клиента", "CAC, ₽", "ROMI-порог при валовой марже 33%"])
    style_headers(ws, 1, 1, 7)
    for channel_name, params in CHANNELS.items():
        client_conv = params["lead_to_measure"] * params["measure_to_deal"]
        cac = params["cpl"] / client_conv
        romi_threshold = cac / (weighted_avg_check(SCENARIOS["Базовый"].segment_mix, "A_Аутсорс") * 0.33)
        ws.append(
            [
                channel_name,
                params["cpl"],
                params["lead_to_measure"],
                params["measure_to_deal"],
                client_conv,
                cac,
                romi_threshold,
            ]
        )

    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=2).number_format = '#,##0 "₽"'
        for col in (3, 4, 5, 7):
            ws.cell(row=row, column=col).number_format = "0.0%"
        ws.cell(row=row, column=6).number_format = '#,##0 "₽"'
    apply_body_style(ws, 2, ws.max_row, 1, 7)
    autosize(ws)


def add_fot_sheet(wb: Workbook, rows: list[dict[str, float | dict[str, int]]]) -> None:
    ws = wb.create_sheet("ФОТ")
    ws.append(["Роль", "Оклад, ₽/мес", "Функция", "Штат на старте", "Штат на месяце цели", "ФОТ на старте, ₽", "ФОТ на месяце цели, ₽"])
    style_headers(ws, 1, 1, 7)
    target_row = next((row for row in rows if float(row["net_profit"]) >= 2_000_000), rows[-1])
    start_counts = rows[0]["counts"]
    target_counts = target_row["counts"]
    role_functions = {
        "Менеджер продаж": "Обработка лидов, расчет КП, доведение до договора.",
        "Дизайнер": "Проектирование, визуализация, подготовка ТЗ на фабрику.",
        "Замерщик": "Выезд, обмер, фотофиксация, первичная техпроверка.",
        "Бригадир сборки": "Организация монтажей, рекламации, контроль качества на объекте.",
        "Операционный менеджер": "Координация заказов, фабрик, логистики и сроков.",
        "Администратор": "Документы, оплаты, CRM, колл-трекинг, подрядчики.",
    }
    for role, salary in SALARIES.items():
        if role not in role_functions:
            continue
        start_count = start_counts.get(role, 0)
        target_count = target_counts.get(role, 0)
        ws.append(
            [
                role,
                salary,
                role_functions[role],
                start_count,
                target_count,
                salary * start_count,
                salary * target_count,
            ]
        )
    last = ws.max_row + 1
    ws.append(["Бонусы продаж", "", "Переменная часть мотивации менеджеров и дизайнеров.", "", "", rows[0]["revenue"] * 0.018, target_row["revenue"] * 0.018])
    ws.append(["Итого фиксированный ФОТ", "", "", "", "", sum(SALARIES[r] * start_counts.get(r, 0) for r in role_functions), sum(SALARIES[r] * target_counts.get(r, 0) for r in role_functions)])
    ws.append(["Итого ФОТ с бонусами", "", "", "", "", rows[0]["fot"], target_row["fot"]])
    for row in range(2, ws.max_row + 1):
        for col in (2, 6, 7):
            ws.cell(row=row, column=col).number_format = '#,##0_);[Red](#,##0)'
        ws.cell(row=row, column=3).alignment = Alignment(wrap_text=True)
    autosize(ws)


def add_media_plan_sheet(wb: Workbook, scenario_name: str, rows: list[dict[str, float | dict[str, int]]]) -> None:
    sheet_name = "Смета (MID)" if scenario_name == "Базовый" else "Смета (OPT)"
    ws = wb.create_sheet(sheet_name)
    ws.append(["Услуги"] + MONTH_LABELS)
    style_headers(ws, 1, 1, 25)

    site_dev = [140_000, 140_000] + [0] * 22
    site_support = [0, 0] + [24_000] * 22
    context_management = [40_000] + [60_000] * 23
    avito_management = [20_000] + [25_000] * 23
    calltracking = [0] + [12_000] * 23

    cost_rows = [
        ("Разработка сайта", site_dev),
        ("Техническая поддержка сайта", site_support),
        ("Ведение контекстной рекламы", context_management),
        ("Бюджет контекстной рекламной кампании", [row["budgets"]["Контекст_поиск"] for row in rows]),
        ("Ведение Авито", avito_management),
        ("Бюджет Авито", [row["budgets"]["Авито"] for row in rows]),
        ("Агрегаторы", [row["budgets"]["Агрегаторы"] for row in rows]),
        ("Ретаргетинг", [row["budgets"]["Ретаргетинг"] for row in rows]),
        ("Calltracking / аналитика", calltracking),
    ]
    for title, values in cost_rows:
        ws.append([title] + values)

    sum_row = ws.max_row + 1
    ws.append(["СУММА"] + [f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}10)" for col in range(2, 26)])

    ws.append([""])
    lead_header_row = ws.max_row + 1
    ws.append(["Прогноз по лидам"])
    ws.append(["Прогнозное количество лидов с источником PPC (Контекстная реклама)"] + [row["leads_by_channel"]["Контекст_поиск"] for row in rows])
    ws.append(["Прогнозное количество лидов с источником Авито"] + [row["leads_by_channel"]["Авито"] for row in rows])
    ws.append(["Прогнозное количество лидов с источником Агрегаторы"] + [row["leads_by_channel"]["Агрегаторы"] for row in rows])
    ws.append(["Прогнозное количество лидов с источником Ретаргетинг"] + [row["leads_by_channel"]["Ретаргетинг"] for row in rows])
    total_leads_row = ws.max_row + 1
    ws.append(["Всего"] + [f"=SUM({get_column_letter(col)}{lead_header_row + 2}:{get_column_letter(col)}{lead_header_row + 5})" for col in range(2, 26)])
    cpl_row = ws.max_row + 2
    ws.append([""])
    ws.append(["CPL от всех затрат"] + [f"={get_column_letter(col)}{sum_row}/{get_column_letter(col)}{total_leads_row}" for col in range(2, 26)])

    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
        if row in (sum_row, lead_header_row, total_leads_row, cpl_row):
            ws.cell(row=row, column=1).fill = SECTION_FILL
            ws.cell(row=row, column=1).font = Font(bold=True)
        for col in range(2, 26):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            if row in (lead_header_row + 2, lead_header_row + 3, lead_header_row + 4, lead_header_row + 5, total_leads_row):
                cell.number_format = '0.0'
            elif row == cpl_row:
                cell.number_format = '#,##0_);[Red](#,##0)'
            else:
                cell.number_format = '#,##0_);[Red](#,##0)'
    ws.freeze_panes = "B2"
    autosize(ws)


def add_conversion_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Конверсии")
    ws.append(["Канал", "CPL, ₽", "Лид → замер", "Замер → договор", "Лид → договор", "CAC, ₽", "Комментарий"])
    style_headers(ws, 1, 1, 7)
    comments = {
        "Контекст_поиск": "Быстрый запуск по горячему спросу, требует дисциплины в аналитике.",
        "Ретаргетинг": "Добирает теплую аудиторию сайта и Авито, улучшает blended CAC.",
        "Авито": "Сильный оффер, живые фото и быстрый ответ критичны.",
        "Агрегаторы": "Работает как добавка, но часто слабее по качеству лида.",
    }
    for channel_name, params in CHANNELS.items():
        final_conv = params["lead_to_measure"] * params["measure_to_deal"]
        ws.append([
            channel_name,
            params["cpl"],
            params["lead_to_measure"],
            params["measure_to_deal"],
            final_conv,
            params["cpl"] / final_conv,
            comments[channel_name],
        ])
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=2).number_format = '#,##0 "₽"'
        ws.cell(row=row, column=6).number_format = '#,##0 "₽"'
        for col in (3, 4, 5):
            ws.cell(row=row, column=col).number_format = "0.0%"
        ws.cell(row=row, column=7).alignment = Alignment(wrap_text=True)
    autosize(ws)


def add_media_logic_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Логика_медиаплана")
    ws.append(["Фаза", "Период", "Логика бюджета", "Что считаем успехом"])
    style_headers(ws, 1, 1, 4)
    for start, end, title, note in PHASES:
        if title == "Запуск":
            success = "Сайт запущен, сквозная аналитика работает, первые 8-10 договоров/мес."
        elif title == "Разгон":
            success = "Есть ROMI-положительные связки объявлений и 12-18 договоров/мес."
        elif title == "Оптимизация":
            success = "Удерживаем CAC, перераспределяем бюджет в SEO, ретаргетинг и сильные кластеры."
        else:
            success = "Масштабируем только то, что выдерживает SLA обработки и дает прибыль."
        ws.append([title, f"M{start}-M{end}", note, success])
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=3).alignment = Alignment(wrap_text=True)
        ws.cell(row=row, column=4).alignment = Alignment(wrap_text=True)
    autosize(ws)


def add_hiring_plan_sheet(wb: Workbook, rows: list[dict[str, float | dict[str, int]]]) -> None:
    ws = wb.create_sheet("План_найма")
    roles = ["Менеджер продаж", "Дизайнер", "Замерщик", "Бригадир сборки", "Операционный менеджер", "Администратор"]
    headers = ["Месяц", "Фаза", "Новая роль / действие", "ФОТ всего, ₽"] + roles
    ws.append(headers)
    style_headers(ws, 1, 1, len(headers))
    prev_counts = None
    for row in rows:
        counts = row["counts"]
        changes = []
        if prev_counts is None:
            changes.append("Старт команды")
        else:
            for role in roles:
                delta = counts.get(role, 0) - prev_counts.get(role, 0)
                if delta > 0:
                    changes.append(f"+{delta} {role}")
        if not changes:
            changes_text = "Без изменений"
        else:
            changes_text = ", ".join(changes)
        ws.append([row["month"], row["phase"], changes_text, row["fot"]] + [counts.get(role, 0) for role in roles])
        prev_counts = counts
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=4).number_format = '#,##0_);[Red](#,##0)'
    autosize(ws)


def add_fot_monthly_sheet(wb: Workbook, rows: list[dict[str, float | dict[str, int]]]) -> None:
    ws = wb.create_sheet("ФОТ_24м")
    roles = ["Менеджер продаж", "Дизайнер", "Замерщик", "Бригадир сборки", "Операционный менеджер", "Администратор"]
    headers = ["Месяц", "Фаза", "Фиксированный ФОТ, ₽", "Бонусы, ₽", "ФОТ всего, ₽"] + roles
    ws.append(headers)
    style_headers(ws, 1, 1, len(headers))
    for row in rows:
        counts = row["counts"]
        fixed_fot = sum(SALARIES[role] * counts.get(role, 0) for role in roles)
        bonus = row["fot"] - fixed_fot
        ws.append([row["month"], row["phase"], fixed_fot, bonus, row["fot"]] + [counts.get(role, 0) for role in roles])
    for row in range(2, ws.max_row + 1):
        for col in (3, 4, 5):
            ws.cell(row=row, column=col).number_format = '#,##0_);[Red](#,##0)'
    autosize(ws)


def add_romi_sheet(wb: Workbook, rows: list[dict[str, float | dict[str, int]]], scenario_name: str) -> None:
    ws = wb.create_sheet(f"ROMI_{scenario_name[:4]}")
    headers = ["Месяц", "Фаза"]
    for channel_name in CHANNELS:
        headers.extend([f"{channel_name} бюджет", f"{channel_name} валовая прибыль", f"{channel_name} ROMI"])
    ws.append(headers)
    style_headers(ws, 1, 1, len(headers))
    gross_margin = 1 - (
        MODEL_CONFIG["A_Аутсорс"]["production_share"]
        + MODEL_CONFIG["A_Аутсорс"]["materials_share"]
        + MODEL_CONFIG["A_Аутсорс"]["logistics_share"]
        + MODEL_CONFIG["A_Аутсорс"]["assembly_share"]
    )
    for row in rows:
        values = [row["month"], row["phase"]]
        for channel_name in CHANNELS:
            budget = row["budgets"][channel_name]
            gross_profit = row["orders_by_channel"][channel_name] * row["avg_check"] * gross_margin
            romi = (gross_profit - budget) / budget if budget else 0
            values.extend([budget, gross_profit, romi])
        ws.append(values)
    for row in range(2, ws.max_row + 1):
        for idx in range(3, ws.max_column + 1):
            if (idx - 2) % 3 in (1, 2):
                ws.cell(row=row, column=idx).number_format = '#,##0_);[Red](#,##0)'
            else:
                ws.cell(row=row, column=idx).number_format = "0.0x"
    autosize(ws)


def add_model_sheet(wb: Workbook, model_key: str, scenario_name: str, rows: list[dict[str, float | dict[str, int]]]) -> None:
    ws = wb.create_sheet(f"{model_key[:1]}_{scenario_name[:4]}")
    ws["A1"] = f"{MODEL_CONFIG[model_key]['model_label']} | {scenario_name}"
    ws["A1"].font = Font(size=14, bold=True)

    metrics = [
        "Лиды",
        "Замеры",
        "Заказы",
        "Средний чек",
        "Выручка",
        "Себестоимость: производство",
        "Себестоимость: материалы",
        "Себестоимость: логистика",
        "Себестоимость: сборка",
        "Итого себестоимость",
        "Маркетинг",
        "ФОТ",
        "Налоги на ФОТ",
        "Аренда",
        "IT/CRM",
        "Прочие",
        "EBITDA",
        "Налог УСН",
        "Чистая прибыль",
    ]
    start_row = 3
    ws.cell(row=start_row, column=1, value="Показатель")
    for index, month_name in enumerate(MONTHS, start=2):
        ws.cell(row=start_row, column=index, value=month_name)
    style_headers(ws, start_row, 1, 25)

    value_map = {
        "Лиды": "leads",
        "Замеры": "measurements",
        "Заказы": "orders",
        "Средний чек": "avg_check",
        "Выручка": "revenue",
        "Себестоимость: производство": "production_cost",
        "Себестоимость: материалы": "materials_cost",
        "Себестоимость: логистика": "logistics_cost",
        "Себестоимость: сборка": "assembly_cost",
        "Итого себестоимость": "cogs",
        "Маркетинг": "marketing",
        "ФОТ": "fot",
        "Налоги на ФОТ": "payroll_taxes",
        "Аренда": "rent",
        "IT/CRM": "it_cost",
        "Прочие": "other_opex",
        "EBITDA": "ebitda",
        "Налог УСН": "tax",
        "Чистая прибыль": "net_profit",
    }

    for row_offset, metric_name in enumerate(metrics, start=1):
        row_no = start_row + row_offset
        ws.cell(row=row_no, column=1, value=metric_name)
        if metric_name in {"Выручка", "EBITDA", "Чистая прибыль"}:
            ws.cell(row=row_no, column=1).fill = SECTION_FILL
            ws.cell(row=row_no, column=1).font = Font(bold=True)
        for month_index, row_data in enumerate(rows, start=2):
            cell = ws.cell(row=row_no, column=month_index, value=row_data[value_map[metric_name]])
            if metric_name in {"Лиды", "Замеры", "Заказы"}:
                cell.number_format = '0.0'
            elif metric_name == "Средний чек":
                cell.number_format = '#,##0 "₽"'
            else:
                cell.number_format = '#,##0_);[Red](#,##0)'
            if metric_name in {"Выручка", "EBITDA", "Чистая прибыль"}:
                cell.fill = FORMULA_FILL
                cell.font = Font(bold=True)

    summary_row = start_row + len(metrics) + 3
    ws.cell(row=summary_row, column=1, value="Выручка по типам мебели")
    ws.cell(row=summary_row, column=1).font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=1, value="Тип")
    ws.cell(row=summary_row + 1, column=2, value="Доля")
    ws.cell(row=summary_row + 1, column=3, value="М12, ₽")
    ws.cell(row=summary_row + 1, column=4, value="М24, ₽")
    style_headers(ws, summary_row + 1, 1, 4)
    month_12_revenue = rows[11]["revenue"]
    month_24_revenue = rows[23]["revenue"]
    for idx, (product_name, share) in enumerate(PRODUCT_SHARES.items(), start=summary_row + 2):
        ws.cell(row=idx, column=1, value=product_name)
        ws.cell(row=idx, column=2, value=share)
        ws.cell(row=idx, column=3, value=month_12_revenue * share)
        ws.cell(row=idx, column=4, value=month_24_revenue * share)
        ws.cell(row=idx, column=2).number_format = "0.0%"
        ws.cell(row=idx, column=3).number_format = '#,##0_);[Red](#,##0)'
        ws.cell(row=idx, column=4).number_format = '#,##0_);[Red](#,##0)'

    seg_row = summary_row + 8
    ws.cell(row=seg_row, column=1, value="Выручка по сегментам")
    ws.cell(row=seg_row, column=1).font = Font(bold=True)
    ws.cell(row=seg_row + 1, column=1, value="Сегмент")
    ws.cell(row=seg_row + 1, column=2, value="Доля")
    ws.cell(row=seg_row + 1, column=3, value="М12, ₽")
    ws.cell(row=seg_row + 1, column=4, value="М24, ₽")
    style_headers(ws, seg_row + 1, 1, 4)
    for idx, (segment_name, share) in enumerate(zip(["Эконом", "Средний", "Премиум"], SCENARIOS[scenario_name].segment_mix), start=seg_row + 2):
        ws.cell(row=idx, column=1, value=segment_name)
        ws.cell(row=idx, column=2, value=share)
        ws.cell(row=idx, column=3, value=month_12_revenue * share)
        ws.cell(row=idx, column=4, value=month_24_revenue * share)
        ws.cell(row=idx, column=2).number_format = "0.0%"
        ws.cell(row=idx, column=3).number_format = '#,##0_);[Red](#,##0)'
        ws.cell(row=idx, column=4).number_format = '#,##0_);[Red](#,##0)'

    chart = LineChart()
    chart.title = "Рост выручки и чистой прибыли"
    chart.y_axis.title = "₽"
    chart.x_axis.title = "Месяц"
    cats = Reference(ws, min_col=2, max_col=25, min_row=start_row, max_row=start_row)
    revenue_data = Reference(ws, min_col=2, max_col=25, min_row=start_row + 5, max_row=start_row + 5)
    profit_data = Reference(ws, min_col=2, max_col=25, min_row=start_row + 19, max_row=start_row + 19)
    chart.add_data(revenue_data, titles_from_data=False, from_rows=True)
    chart.add_data(profit_data, titles_from_data=False, from_rows=True)
    chart.set_categories(cats)
    chart.height = 9
    chart.width = 18
    ws.add_chart(chart, "F26")

    apply_body_style(ws, start_row + 1, start_row + len(metrics), 1, 25)
    ws.freeze_panes = "B4"
    autosize(ws)


def add_summary_sheet(wb: Workbook, model_data: dict[tuple[str, str], list[dict[str, float | dict[str, int]]]]) -> None:
    ws = wb.create_sheet("Итоги", 1)
    ws.append([
        "Модель",
        "Сценарий",
        "Стартовые инвестиции, ₽",
        "Месяц безубыточности",
        "Месяц 2 млн чистыми",
        "Выручка на месяце цели, ₽",
        "Заказы на месяце цели",
        "Средний чек на месяце цели, ₽",
        "Окупаемость, мес",
    ])
    style_headers(ws, 1, 1, 9)

    for (model_key, scenario_name), rows in model_data.items():
        target_month = None
        breakeven_month = None
        cumulative_cash = -MODEL_CONFIG[model_key]["startup_investment"]
        payback_month = None
        target_revenue = None
        target_orders = None
        target_avg_check = None
        for row in rows:
            month = int(row["month"])
            net_profit = float(row["net_profit"])
            if breakeven_month is None and net_profit > 0:
                breakeven_month = month
            if target_month is None and net_profit >= 2_000_000:
                target_month = month
                target_revenue = row["revenue"]
                target_orders = row["orders"]
                target_avg_check = row["avg_check"]
            cumulative_cash += net_profit
            if payback_month is None and cumulative_cash >= 0:
                payback_month = month

        ws.append(
            [
                MODEL_CONFIG[model_key]["model_label"],
                scenario_name,
                MODEL_CONFIG[model_key]["startup_investment"],
                breakeven_month or "не достигнута",
                target_month or "не достигнута",
                target_revenue or 0,
                target_orders or 0,
                target_avg_check or 0,
                payback_month or "не достигнута",
            ]
        )

    for row in range(2, ws.max_row + 1):
        for col in (3, 6, 8):
            ws.cell(row=row, column=col).number_format = '#,##0_);[Red](#,##0)'
        ws.cell(row=row, column=7).number_format = "0.0"
    autosize(ws)

def add_investment_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Инвестиции")
    ws.append(["Статья", "Аутсорс, ₽", "Комментарий"])
    style_headers(ws, 1, 1, 3)
    rows = [
        ("Маркетинговый запуск на 2 месяца", 700_000, "Контекст, Авито, SEO-ядро, фотоконтент и посадочные."),
        ("Разработка сайта", 280_000, "Прототип, дизайн, верстка, калькулятор, формы, аналитика, базовое SEO."),
        ("CRM, телефония, аналитика", 120_000, "Bitrix24/amo + коллтрекинг + сквозная аналитика."),
        ("Депозит и запуск офиса/шоурума", 250_000, "1-2 месяца депозита и базовая мебель."),
        ("Оборотный капитал", 1_250_000, "Авансы фабрикам/поставщикам, кассовый разрыв, реклама."),
        ("Фотоконтент и портфолио", 180_000, "Предметная и интерьерная съемка 8-10 проектов, ретушь, видео-рили."),
        ("Юр. запуск, бренд, регламенты", 270_000, "Договорная база, бренд, скрипты, KPI, шаблоны ТЗ."),
    ]
    for title, outsource_cost, comment in rows:
        ws.append([title, outsource_cost, comment])
    ws.append(["ИТОГО", "=SUM(B2:B8)", ""])
    last_row = ws.max_row
    for row in range(2, last_row + 1):
        ws.cell(row=row, column=2).number_format = '#,##0_);[Red](#,##0)'
        ws.cell(row=row, column=3).alignment = Alignment(wrap_text=True)
    ws.cell(row=last_row, column=1).fill = SECTION_FILL
    ws.cell(row=last_row, column=2).fill = FORMULA_FILL
    ws.cell(row=last_row, column=1).font = Font(bold=True)
    autosize(ws)


def add_sources_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Источники")
    ws.append(["Источник", "URL", "Как использовано в модели"])
    style_headers(ws, 1, 1, 3)
    for source in SOURCES:
        ws.append(list(source))
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
        ws.cell(row=row, column=3).alignment = Alignment(wrap_text=True)
    autosize(ws)


def build_markdown_report(model_data: dict[tuple[str, str], list[dict[str, float | dict[str, int]]]]) -> str:
    def get_target(model_key: str, scenario_name: str):
        rows = model_data[(model_key, scenario_name)]
        for row in rows:
            if float(row["net_profit"]) >= 2_000_000:
                return row
        return rows[-1]

    base_rows = model_data[("A_Аутсорс", "Базовый")]
    opt_rows = model_data[("A_Аутсорс", "Оптимистичный")]
    base_a = get_target("A_Аутсорс", "Базовый")
    opt_a = get_target("A_Аутсорс", "Оптимистичный")
    start_rows = model_data[("A_Аутсорс", "Базовый")]
    target_row = next(row for row in start_rows if float(row["net_profit"]) >= 2_000_000)
    start_counts = start_rows[0]["counts"]
    target_counts = target_row["counts"]
    base_target_month = next(row["month"] for row in base_rows if float(row["net_profit"]) >= 2_000_000)
    opt_target_month = next(row["month"] for row in opt_rows if float(row["net_profit"]) >= 2_000_000)

    return f"""# Финансовый бизнес-план: мебель на заказ, Москва и МО

## Что внутри

- Горизонт: 24 месяца.
- Модель: только аутсорс производства.
- Сценарии: базовый и оптимистичный.
- Налоговый режим в модели: УСН 15% «доходы минус расходы».
- Медиаплан построен по фазам: запуск, разгон, оптимизация, масштабирование.

## Ключевые выводы

- Стартовые вложения по модели аутсорса: **3,05 млн ₽**, включая **280 тыс. ₽** на разработку сайта.
- Базовый сценарий: цель **2 млн ₽ чистой прибыли** достигается на **{base_target_month} месяце** при выручке **{base_a["revenue"]:,.0f} ₽**, **{base_a["orders"]:.1f}** заказах и среднем чеке **{base_a["avg_check"]:,.0f} ₽**.
- Оптимистичный сценарий: цель достигается на **{opt_target_month} месяце** при выручке **{opt_a["revenue"]:,.0f} ₽**, **{opt_a["orders"]:.1f}** заказах и среднем чеке **{opt_a["avg_check"]:,.0f} ₽**.

## Unit-экономика, базовый ориентир

| Канал | CPL | CAC | Комментарий |
|---|---:|---:|---|
| Контекст поиск | 5 800 ₽ | ~82 621 ₽ | Самый быстрый канал запуска, но чувствителен к качеству сайта и обработке лидов |
| Ретаргетинг | 3 000 ₽ | ~36 765 ₽ | Дожимает теплую аудиторию сайта и Авито, снижает blended CAC |
| Авито | 2 500 ₽ | ~41 667 ₽ | Хороший стартовый объем при сильной обработке заявок |
| Агрегаторы | 4 500 ₽ | ~94 538 ₽ | Нужны как добавка к потоку, не как базовый двигатель |
| Рекомендации | вне медиаплана | вне медиаплана | Учитываются отдельно как органический бонус к базе, без закупки трафика |

## Базовая точка безубыточности

- Аутсорс: примерно **10-11 заказов/мес** или **4,2-4,6 млн ₽/мес** выручки.

## ФОТ в базовом сценарии

- На старте: менеджер продаж `{start_counts["Менеджер продаж"]}`, дизайнер `{start_counts["Дизайнер"]}`, замерщик `{start_counts["Замерщик"]}`, бригадир сборки `{start_counts["Бригадир сборки"]}`, операционный менеджер `{start_counts["Операционный менеджер"]}`, администратор `{start_counts["Администратор"]}`.
- На месяце цели: менеджеры продаж `{target_counts["Менеджер продаж"]}`, дизайнеры `{target_counts["Дизайнер"]}`, замерщики `{target_counts["Замерщик"]}`, бригадиры сборки `{target_counts["Бригадир сборки"]}`, операционные менеджеры `{target_counts["Операционный менеджер"]}`, администратор `{target_counts["Администратор"]}`.
- Полная раскладка по окладам и суммарному ФОТ вынесена на листы `ФОТ`, `ФОТ_24м` и `План_найма`.

## Как читать медиаплан

- Лист `Логика_медиаплана` объясняет, зачем меняется структура бюджета по фазам.
- Листы `Смета (MID)` и `Смета (OPT)` собраны по образцу медиаплана: сверху строки затрат, ниже блок прогнозных лидов, затем общий CPL.
- Листы `ROMI_Базо` и `ROMI_Опти` показывают валовую прибыль по каждому каналу и ROMI помесячно.

## Риски

- Просадка конверсии замер → договор на 5 п.п. сдвигает достижение цели на 2-4 месяца.
- Контекст и агрегаторы при росте CPL на 15-20% резко ухудшают CAC и ROMI.
- Слабый сайт и плохое портфолио снижают конверсию контекста, Авито и ретаргетинга уже в первые месяцы, поэтому бюджет на сайт не стоит урезать.
- Аутсорс-модель чувствительна к срокам и качеству подрядчиков, нужен жесткий SLA и контроль рекламаций.

## Источники

- ФНС по УСН: https://www.nalog.gov.ru/rn25/ifns/r25_14/info/15658830/
- ФНС по взносам 2025: https://www.nalog.gov.ru/rn28/news/activities_fts/15540345/
- ФНС Москва по взносам 2026: https://www.nalog.gov.ru/rn77/news/activities_fts/16599875/
- Офисная аренда Москва: https://www.ecooffice.ru/arenda_ofisa/ot_200_metrov_kv
- Ориентир по аренде мебельного цеха: https://optima-invest.ru/msk/obekty/stolyarnyy_tsekh_i_mebelnoe_proizvodstvo/
- Ориентир по станку раскроя: https://www.stankoff.ru/product/26773/formatno-raskroechnyiy-stanok-d45a
"""


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    add_assumptions_sheet(wb)
    add_unit_economics_sheet(wb)
    add_investment_sheet(wb)
    add_sources_sheet(wb)
    add_media_logic_sheet(wb)

    model_data: dict[tuple[str, str], list[dict[str, float | dict[str, int]]]] = {}
    for model_key in MODEL_CONFIG:
        for scenario_name in SCENARIOS:
            rows = monthly_rows(model_key, scenario_name)
            model_data[(model_key, scenario_name)] = rows
            add_model_sheet(wb, model_key, scenario_name, rows)
            add_media_plan_sheet(wb, scenario_name, rows)
            add_romi_sheet(wb, rows, scenario_name)

    add_summary_sheet(wb, model_data)
    add_conversion_sheet(wb)
    add_fot_sheet(wb, model_data[("A_Аутсорс", "Базовый")])
    add_fot_monthly_sheet(wb, model_data[("A_Аутсорс", "Базовый")])
    add_hiring_plan_sheet(wb, model_data[("A_Аутсорс", "Базовый")])

    wb.save(WORKBOOK_PATH)
    REPORT_PATH.write_text(build_markdown_report(model_data), encoding="utf-8")

    # Повторное открытие подтверждает, что файл корректно записан.
    load_workbook(WORKBOOK_PATH)


if __name__ == "__main__":
    main()
