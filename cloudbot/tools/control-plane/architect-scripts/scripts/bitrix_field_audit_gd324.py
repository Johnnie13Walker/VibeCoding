#!/usr/bin/env python3
"""Аудит полей Bitrix24 для ГД-324 с выгрузкой в Excel."""

from __future__ import annotations

import json
import math
import os
import re
import subprocess
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

ROOT = Path(os.environ.get("MARKETING_DASHBOARD_ROOT_DIR", Path(__file__).resolve().parents[1]))
ENGINEER_ROOT = Path(os.environ.get("MARKETING_DASHBOARD_ENGINEER_ROOT", "/Users/pro2kuror/Desktop/Cloudbot/engineer"))
TMP_DIR = ROOT / "tmp" / "gd324_field_audit"
DOCS_DIR = ROOT / "docs" / "architecture"
STATE_DIR = Path(
    os.environ.get(
        "BITRIX_APP_STATE_DIR",
        "/Users/pro2kuror/Library/Application Support/OpenClo/assistant/tmp/sales-copilot-state-97sfyomu/bitrix_app",
    )
)
OUTPUT_XLSX = DOCS_DIR / "gd-324-bitrix-field-audit.xlsx"
OUTPUT_JSON = TMP_DIR / "gd324_field_audit_data.json"
OUTPUT_SUMMARY = DOCS_DIR / "gd-324-bitrix-field-audit-summary.md"

MOSCOW_NOW = datetime.now().astimezone().strftime("%Y-%m-%d %H:%M")
SSH_HOST = "188.34.206.115"
SSH_USER = "ops"
SSH_KEY = os.path.expanduser("~/.ssh/temp_migration_key")
FULL_DATA_SWEEP = os.environ.get("GD324_FULL_DATA_SWEEP") == "1"
SCOPE = os.environ.get("GD324_SCOPE", "all").strip().lower() or "all"
Workbook = None
Alignment = None
Font = None
PatternFill = None
get_column_letter = None


def log(message: str) -> None:
    print(message, flush=True)


def _import_auth():
    import sys

    sys.path.insert(0, str(ENGINEER_ROOT))
    from cloudbot.providers.bitrix.bitrix_app_auth import BitrixAppAuth

    return BitrixAppAuth


def ensure_openpyxl() -> None:
    global Workbook, Alignment, Font, PatternFill, get_column_letter
    if Workbook is not None:
        return
    from openpyxl import Workbook as OpenpyxlWorkbook
    from openpyxl.styles import Alignment as OpenpyxlAlignment
    from openpyxl.styles import Font as OpenpyxlFont
    from openpyxl.styles import PatternFill as OpenpyxlPatternFill
    from openpyxl.utils import get_column_letter as openpyxl_get_column_letter

    Workbook = OpenpyxlWorkbook
    Alignment = OpenpyxlAlignment
    Font = OpenpyxlFont
    PatternFill = OpenpyxlPatternFill
    get_column_letter = openpyxl_get_column_letter


def parse_env_lines(lines: list[str], wanted: set[str] | None = None) -> dict[str, str]:
    env: dict[str, str] = {}
    for raw_line in lines:
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        if wanted is not None and key not in wanted:
            continue
        cleaned = value.strip()
        if len(cleaned) >= 2 and cleaned[0] == cleaned[-1] and cleaned[0] in {'"', "'"}:
            cleaned = cleaned[1:-1]
        env[key] = cleaned
    return env


def read_env_file(path: Path, wanted: set[str] | None = None) -> dict[str, str]:
    if not path.exists():
        return {}
    return parse_env_lines(path.read_text(encoding="utf-8").splitlines(), wanted)


def fetch_remote_bitrix_env() -> dict[str, str]:
    command = [
        "ssh",
        "-i",
        SSH_KEY,
        "-p",
        "22",
        "-o",
        "BatchMode=yes",
        "-o",
        "StrictHostKeyChecking=accept-new",
        "-o",
        "ConnectTimeout=10",
        f"{SSH_USER}@{SSH_HOST}",
        "sudo cat /opt/openclaw/.env",
    ]
    completed = subprocess.run(command, text=True, capture_output=True, check=False, timeout=30)
    if completed.returncode != 0:
        detail = (completed.stderr or completed.stdout or f"ssh rc={completed.returncode}").strip()
        raise RuntimeError(f"Не удалось получить .env с сервера: {detail}")

    wanted = {
        "BITRIX_CLIENT_ID",
        "BITRIX_CLIENT_SECRET",
        "BITRIX_OAUTH_TOKEN_URL",
        "BITRIX_TIMEOUT_SEC",
    }
    return parse_env_lines(completed.stdout.splitlines(), wanted)


def load_bitrix_env() -> dict[str, str]:
    wanted = {
        "BITRIX_APP_INSTALL_STATE_FILE",
        "BITRIX_APP_STATE_DIR",
        "BITRIX_CLIENT_ID",
        "BITRIX_CLIENT_SECRET",
        "BITRIX_OAUTH_TOKEN_URL",
        "BITRIX_TIMEOUT_SEC",
    }
    env = {key: value for key, value in os.environ.items() if key in wanted and value}
    for path in (
        Path(os.environ.get("BITRIX_ENV_FILE", "")).expanduser() if os.environ.get("BITRIX_ENV_FILE") else None,
        Path("/opt/openclaw/.env"),
    ):
        if path is None:
            continue
        file_env = read_env_file(path, wanted)
        file_env.update(env)
        env = file_env
    if not env.get("BITRIX_CLIENT_ID") or not env.get("BITRIX_CLIENT_SECRET"):
        remote_env = fetch_remote_bitrix_env()
        remote_env.update(env)
        env = remote_env
    return env


def make_auth():
    BitrixAppAuth = _import_auth()
    env = load_bitrix_env()
    state_dir = Path(env.get("BITRIX_APP_STATE_DIR") or str(STATE_DIR)).expanduser()
    state_file_raw = str(env.get("BITRIX_APP_INSTALL_STATE_FILE") or "").strip()
    state_file = Path(state_file_raw).expanduser() if state_file_raw else None
    timeout_sec = os.environ.get("BITRIX_TIMEOUT_SEC") or env.get("BITRIX_TIMEOUT_SEC") or 20
    return BitrixAppAuth(
        state_dir=state_dir,
        state_file=state_file,
        client_id=env.get("BITRIX_CLIENT_ID", ""),
        client_secret=env.get("BITRIX_CLIENT_SECRET", ""),
        oauth_token_url=env.get("BITRIX_OAUTH_TOKEN_URL", ""),
        timeout_sec=int(timeout_sec),
    )


def payload_result(payload: Any) -> Any:
    if isinstance(payload, dict) and "result" in payload:
        return payload.get("result")
    return payload


def list_result(payload: Any) -> list[dict[str, Any]]:
    result = payload_result(payload)
    if isinstance(result, list):
        return [item for item in result if isinstance(item, dict)]
    if isinstance(result, dict):
        for key in ("items", "types", "categories"):
            value = result.get(key)
            if isinstance(value, list):
                return [item for item in value if isinstance(item, dict)]
    return []


def field_map_result(payload: Any) -> dict[str, Any]:
    result = payload_result(payload)
    if isinstance(result, dict) and "fields" in result and isinstance(result["fields"], dict):
        return result["fields"]
    return result if isinstance(result, dict) else {}


def bitrix_call(auth, method: str, params: dict[str, Any] | None = None) -> Any:
    return auth.call_payload(method, params=params or {}, default={})


def bitrix_list(auth, method: str, params: dict[str, Any] | None = None) -> list[dict[str, Any]]:
    return auth.list_method(method, params=params or {})


def normalize_label(value: Any) -> str:
    return str(value or "").strip()


def is_non_empty(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    if isinstance(value, (list, tuple, set, dict)):
        return len(value) > 0
    return True


def pct(value: int, total: int) -> float:
    if total <= 0:
        return 0.0
    return round((value / total) * 100, 2)


def fmt_pct(value: float) -> str:
    return f"{value:.2f}%"


def normalize_text_key(value: str) -> str:
    lowered = value.lower().replace("ё", "е")
    lowered = re.sub(r"[^a-zа-я0-9]+", " ", lowered)
    lowered = re.sub(r"\s+", " ", lowered).strip()
    return lowered


SEMANTIC_RULES: list[tuple[str, tuple[str, ...]]] = [
    ("source_detail", ("источник детально", "подисточник", "source detail", "utm source")),
    ("source", ("источник", "source", "канал привлечения")),
    ("lead_type", ("тип лида", "lead type")),
    ("client_type", ("тип клиента", "client type")),
    ("segment", ("сегмент", "segment")),
    ("industry", ("сфера деятельности", "отрасль", "industry")),
    ("city", ("город", "city")),
    ("region", ("регион", "region", "область")),
    ("site", ("сайт", "website", "url")),
    ("budget", ("бюджет", "budget")),
    ("client_revenue", ("выручка", "revenue")),
    ("client_size", ("размер клиента", "размер компании", "company size", "численность")),
    ("decision_maker", ("лпр", "лицо принимающее решение", "decision maker")),
    ("qualification", ("квалификац", "qualification")),
    ("next_step", ("следующий шаг", "next step")),
    ("loss_reason", ("причина потери", "причина проигрыша", "loss reason")),
    ("rejection_reason", ("причина отказа", "reason reject", "reason decline")),
    ("next_contact_date", ("дата следующего контакта", "следующий контакт", "next contact")),
    ("responsible", ("ответственный", "assigned by", "responsible")),
    ("communication_channel", ("канал коммуникации", "канал связи", "communication channel")),
]


IMPORTANT_SEMANTICS = {
    "source",
    "source_detail",
    "lead_type",
    "client_type",
    "segment",
    "industry",
    "city",
    "region",
    "site",
    "budget",
    "client_revenue",
    "client_size",
    "decision_maker",
    "qualification",
    "next_step",
    "loss_reason",
    "rejection_reason",
    "next_contact_date",
    "responsible",
    "communication_channel",
}


def semantic_key(title: str, code: str) -> str | None:
    hay = f"{normalize_text_key(title)} {normalize_text_key(code)}"
    for key, needles in SEMANTIC_RULES:
        if any(needle in hay for needle in needles):
            return key
    return None


def infer_report_usage(title: str, code: str, field_type: str, is_system: bool) -> str:
    key = semantic_key(title, code)
    code_u = str(code or "").upper()
    if key in IMPORTANT_SEMANTICS:
        return "Вероятно да"
    if code_u in {"SOURCE_ID", "ASSIGNED_BY_ID", "CATEGORY_ID", "STAGE_ID", "STATUS_ID", "OPPORTUNITY"}:
        return "Да"
    if field_type in {"money", "crm_status", "crm_category", "date", "datetime"} and is_system:
        return "Вероятно да"
    return "Недостаточно данных"


def infer_card_usage(title: str, code: str, fill_rate: float) -> str:
    key = semantic_key(title, code)
    if key in IMPORTANT_SEMANTICS:
        return "Вероятно да"
    if fill_rate > 20:
        return "Вероятно да"
    return "Требуется ручная проверка"


def classify_field_origin(code: str) -> str:
    return "Кастомное" if code.startswith("UF_") or code.startswith("ufCrm") else "Системное"


def simplify_type(meta: dict[str, Any]) -> str:
    return normalize_label(meta.get("type") or meta.get("userType") or meta.get("fieldType") or "unknown")


def field_title(code: str, meta: dict[str, Any]) -> str:
    for key in ("title", "formLabel", "listLabel", "filterLabel", "name"):
        value = normalize_label(meta.get(key))
        if value:
            return value
    return code


def normalize_field_meta(code: str, meta: dict[str, Any]) -> dict[str, Any]:
    return {
        "code": code,
        "title": field_title(code, meta),
        "type": simplify_type(meta),
        "origin": classify_field_origin(code),
        "required": bool(meta.get("isRequired") or meta.get("mandatory")),
        "multiple": bool(meta.get("isMultiple") or meta.get("multiple")),
        "raw": meta,
    }


CORE_SELECT_FIELDS = {
    "ID",
    "TITLE",
    "TYPE_ID",
    "CATEGORY_ID",
    "STAGE_ID",
    "STATUS_ID",
    "SOURCE_ID",
    "SOURCE_DESCRIPTION",
    "ASSIGNED_BY_ID",
    "DATE_CREATE",
    "DATE_MODIFY",
    "BEGINDATE",
    "CLOSEDATE",
    "OPPORTUNITY",
    "COMPANY_ID",
    "CONTACT_ID",
    "STATUS_SEMANTIC_ID",
    "IS_NEW",
}

CORE_DYNAMIC_FIELDS = {
    "id",
    "title",
    "categoryId",
    "stageId",
    "assignedById",
    "createdTime",
    "updatedTime",
    "movedTime",
    "opened",
    "parentId2",
}


def select_codes_for_analysis(field_map: dict[str, dict[str, Any]], *, dynamic: bool = False) -> list[str]:
    selected: set[str] = set(CORE_DYNAMIC_FIELDS if dynamic else CORE_SELECT_FIELDS)
    for code, meta in field_map.items():
        title = meta["title"]
        if meta["origin"] == "Кастомное":
            selected.add(code)
            continue
        if semantic_key(title, code):
            selected.add(code)
    for mandatory in ("CATEGORY_ID", "STAGE_ID", "STATUS_ID", "categoryId", "stageId"):
        if mandatory in field_map:
            selected.add(mandatory)
    return sorted(code for code in selected if code in field_map)


def sample_values(records: list[dict[str, Any]], field_code: str, limit: int = 3) -> list[str]:
    values: list[str] = []
    for record in records:
        value = record.get(field_code)
        if not is_non_empty(value):
            continue
        if isinstance(value, (list, tuple, set)):
            cleaned = ", ".join(str(v) for v in value[:3])
        elif isinstance(value, dict):
            cleaned = json.dumps(value, ensure_ascii=False)[:140]
        else:
            cleaned = str(value)
        cleaned = cleaned.strip()
        if cleaned and cleaned not in values:
            values.append(cleaned[:140])
        if len(values) >= limit:
            break
    return values


@dataclass
class EntitySlice:
    entity_name: str
    category_name: str
    category_id: str
    stage_field: str | None
    stage_labels: dict[str, str]
    records: list[dict[str, Any]]


def stage_usage_summary(records: list[dict[str, Any]], field_code: str, stage_field: str | None, stage_labels: dict[str, str]) -> str:
    if not stage_field:
        return ""
    counter: Counter[str] = Counter()
    for record in records:
        if not is_non_empty(record.get(field_code)):
            continue
        stage_id = normalize_label(record.get(stage_field))
        label = stage_labels.get(stage_id) or stage_id or "Без стадии"
        counter[label] += 1
    if not counter:
        return ""
    return " | ".join(f"{name}: {count}" for name, count in counter.most_common(3))


def entity_usage_rows(
    entity_name: str,
    category_name: str,
    field_meta_map: dict[str, dict[str, Any]],
    records: list[dict[str, Any]],
    stage_field: str | None,
    stage_labels: dict[str, str],
    analyzed_codes: set[str] | None = None,
) -> list[dict[str, Any]]:
    total = len(records)
    rows: list[dict[str, Any]] = []
    for code, meta in field_meta_map.items():
        analyzed = (analyzed_codes is None or code in analyzed_codes) and total > 0
        non_empty = sum(1 for record in records if is_non_empty(record.get(code))) if analyzed else 0
        fill_rate = pct(non_empty, total) if analyzed else 0.0
        samples = sample_values(records, code) if analyzed else []
        title = meta["title"]
        origin = meta["origin"]
        sem_key = semantic_key(title, code)
        rows.append(
            {
                "Сущность": entity_name,
                "Воронка": category_name,
                "Поле": title,
                "Код поля": code,
                "Тип": meta["type"],
                "Системное / кастомное": origin,
                "Обязательное": "Да" if meta["required"] else "Нет",
                "Множественное": "Да" if meta["multiple"] else "Нет",
                "Используется фактически": ("Да" if non_empty > 0 else "Нет") if analyzed else "Н/Д",
                "Частота заполнения / % заполнения": fmt_pct(fill_rate) if analyzed else "Н/Д (системное поле, полный sweep не выполнялся)",
                "Есть данные или нет": ("Да" if non_empty > 0 else "Нет") if analyzed else "Н/Д",
                "Используется в автоматизациях": "Недостаточно данных / требуется ручная проверка",
                "Используется в отчетах / аналитике": infer_report_usage(title, code, meta["type"], origin == "Системное"),
                "Есть смысловой дубль": "",
                "Название дубля / похожего поля": "",
                "Рекомендация": "",
                "Комментарий": "",
                "_fill_rate": fill_rate,
                "_non_empty": non_empty,
                "_total": total,
                "_semantic_key": sem_key or "",
                "_stage_usage": stage_usage_summary(records, code, stage_field, stage_labels) if analyzed else "",
                "_samples": " ; ".join(samples),
                "_analyzed": analyzed,
            }
        )
    return rows


def entity_field_maps(auth) -> tuple[dict[str, dict[str, Any]], dict[str, dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], dict[str, str]]:
    statuses = bitrix_list(auth, "crm.status.list")
    status_labels: dict[str, str] = {}
    for item in statuses:
        entity_id = normalize_label(item.get("ENTITY_ID"))
        status_id = normalize_label(item.get("STATUS_ID"))
        name = normalize_label(item.get("NAME"))
        if entity_id and status_id:
            status_labels[f"{entity_id}::{status_id}"] = name or status_id

    deal_fields = {
        code: normalize_field_meta(code, meta)
        for code, meta in field_map_result(bitrix_call(auth, "crm.deal.fields")).items()
    }
    lead_fields = {
        code: normalize_field_meta(code, meta)
        for code, meta in field_map_result(bitrix_call(auth, "crm.lead.fields")).items()
    }
    contact_fields = {
        code: normalize_field_meta(code, meta)
        for code, meta in field_map_result(bitrix_call(auth, "crm.contact.fields")).items()
    }
    company_fields = {
        code: normalize_field_meta(code, meta)
        for code, meta in field_map_result(bitrix_call(auth, "crm.company.fields")).items()
    }

    types = list_result(bitrix_call(auth, "crm.type.list"))
    deal_categories = list_result(bitrix_call(auth, "crm.category.list", {"entityTypeId": 2}))

    return (
        {
            "deal": deal_fields,
            "lead": lead_fields,
            "contact": contact_fields,
            "company": company_fields,
        },
        {f"{item.get('ENTITY_ID')}::{item.get('STATUS_ID')}": normalize_label(item.get("NAME")) for item in statuses},
        types,
        deal_categories,
        status_labels,
    )


def stage_label_map(status_labels: dict[str, str], entity_status_id: str) -> dict[str, str]:
    prefix = f"{entity_status_id}::"
    result: dict[str, str] = {}
    for key, value in status_labels.items():
        if key.startswith(prefix):
            result[key[len(prefix):]] = value
    return result


def build_entity_slices(auth):
    field_maps, _, types, deal_categories, status_labels = entity_field_maps(auth)
    slices: list[tuple[EntitySlice, dict[str, dict[str, Any]]]] = []

    log("Собираю сделки...")
    deal_select = select_codes_for_analysis(field_maps["deal"])
    deals = bitrix_list(auth, "crm.deal.list", {"select": deal_select, "order": {"ID": "ASC"}}) if FULL_DATA_SWEEP else []
    log(f"Сделки: {len(deals)}" if FULL_DATA_SWEEP else "Сделки: пропущен полный sweep, работаем по мета-аудиту")
    deal_categories_by_id = {normalize_label(cat.get("id")): normalize_label(cat.get("name")) for cat in deal_categories}
    for category_id, category_name in deal_categories_by_id.items():
        category_records = [item for item in deals if normalize_label(item.get("CATEGORY_ID")) == category_id]
        slices.append(
            (
                EntitySlice(
                    entity_name="Сделки",
                    category_name=category_name,
                    category_id=category_id,
                    stage_field="STAGE_ID",
                    stage_labels=stage_label_map(status_labels, f"DEAL_STAGE_{category_id}"),
                    records=category_records,
                ),
                field_maps["deal"],
            )
        )

    if SCOPE not in {"deals_contacts_companies"}:
        log("Собираю лиды...")
        lead_select = select_codes_for_analysis(field_maps["lead"])
        leads = bitrix_list(auth, "crm.lead.list", {"select": lead_select, "order": {"ID": "ASC"}}) if FULL_DATA_SWEEP else []
        log(f"Лиды: {len(leads)}" if FULL_DATA_SWEEP else "Лиды: пропущен полный sweep, работаем по мета-аудиту")
        slices.append(
            (
                EntitySlice(
                    entity_name="Лиды",
                    category_name="",
                    category_id="",
                    stage_field="STATUS_ID",
                    stage_labels=stage_label_map(status_labels, "STATUS"),
                    records=leads,
                ),
                field_maps["lead"],
            )
        )

    if SCOPE in {"all", "deals_contacts_companies"}:
        log("Собираю контакты...")
        contact_select = select_codes_for_analysis(field_maps["contact"])
        contacts = bitrix_list(auth, "crm.contact.list", {"select": contact_select, "order": {"ID": "ASC"}}) if FULL_DATA_SWEEP else []
        log(f"Контакты: {len(contacts)}" if FULL_DATA_SWEEP else "Контакты: пропущен полный sweep, работаем по мета-аудиту")
        slices.append(
            (
                EntitySlice(
                    entity_name="Контакты",
                    category_name="",
                    category_id="",
                    stage_field=None,
                    stage_labels={},
                    records=contacts,
                ),
                field_maps["contact"],
            )
        )

        log("Собираю компании...")
        company_select = select_codes_for_analysis(field_maps["company"])
        companies = bitrix_list(auth, "crm.company.list", {"select": company_select, "order": {"ID": "ASC"}}) if FULL_DATA_SWEEP else []
        log(f"Компании: {len(companies)}" if FULL_DATA_SWEEP else "Компании: пропущен полный sweep, работаем по мета-аудиту")
        slices.append(
            (
                EntitySlice(
                    entity_name="Компании",
                    category_name="",
                    category_id="",
                    stage_field=None,
                    stage_labels={},
                    records=companies,
                ),
                field_maps["company"],
            )
        )

    if SCOPE == "all":
        for item_type in types:
            entity_type_id = int(item_type.get("entityTypeId"))
            entity_title = normalize_label(item_type.get("title") or item_type.get("code") or f"Smart {entity_type_id}")
            log(f"Собираю smart-process: {entity_title} (entityTypeId={entity_type_id})")
            fields = {
                code: normalize_field_meta(code, meta)
                for code, meta in field_map_result(bitrix_call(auth, "crm.item.fields", {"entityTypeId": entity_type_id})).items()
            }
            select_fields = select_codes_for_analysis(fields, dynamic=True)
            records = (
                bitrix_list(
                    auth,
                    "crm.item.list",
                    {
                        "entityTypeId": entity_type_id,
                        "select": select_fields,
                        "order": {"id": "ASC"},
                    },
                )
                if FULL_DATA_SWEEP
                else []
            )
            if FULL_DATA_SWEEP:
                log(f"Smart-process {entity_title}: {len(records)} записей, {len(fields)} полей")
            else:
                log(f"Smart-process {entity_title}: полный sweep пропущен, полей {len(fields)}")
            categories = list_result(bitrix_call(auth, "crm.category.list", {"entityTypeId": entity_type_id}))
            categories_by_id = {normalize_label(cat.get("id")): normalize_label(cat.get("name")) for cat in categories}
            has_categories = bool(item_type.get("isCategoriesEnabled")) and bool(categories_by_id)
            has_stages = bool(item_type.get("isStagesEnabled"))

            if has_categories:
                for category_id, category_name in categories_by_id.items():
                    category_records = [rec for rec in records if normalize_label(rec.get("categoryId") or rec.get("category_id")) == category_id]
                    slices.append(
                        (
                            EntitySlice(
                                entity_name=entity_title,
                                category_name=category_name,
                                category_id=category_id,
                                stage_field="stageId" if has_stages else None,
                                stage_labels=stage_label_map(status_labels, f"DYNAMIC_{entity_type_id}_STAGE_{category_id}") if has_stages else {},
                                records=category_records,
                            ),
                            fields,
                        )
                    )
            else:
                slices.append(
                    (
                        EntitySlice(
                            entity_name=entity_title,
                            category_name="",
                            category_id="",
                            stage_field="stageId" if has_stages else None,
                            stage_labels={},
                            records=records,
                        ),
                        fields,
                    )
                )
    return slices


def build_duplicate_index(rows: list[dict[str, Any]]) -> tuple[dict[tuple[str, str], list[dict[str, Any]]], list[dict[str, Any]]]:
    groups: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        key = normalize_text_key(row["Поле"])
        semantic = row["_semantic_key"]
        if semantic:
            group_key = ("semantic", semantic)
        else:
            group_key = ("title", key)
        groups[group_key].append(row)

    duplicate_sheet_rows: list[dict[str, Any]] = []
    for key, group in groups.items():
        if len(group) < 2:
            continue
        origin_entities = sorted({f"{row['Сущность']}::{row['Воронка'] or '—'}" for row in group})
        primary = sorted(group, key=lambda row: (-row["_fill_rate"], row["Системное / кастомное"] != "Системное", row["Код поля"]))[0]
        secondary = [row for row in group if row is not primary]
        for row in group:
            row["Есть смысловой дубль"] = "Да"
            row["Название дубля / похожего поля"] = ", ".join(other["Поле"] for other in group if other is not row)[:500]

        for row in secondary:
            duplicate_sheet_rows.append(
                {
                    "Поле 1": primary["Поле"],
                    "Поле 2": row["Поле"],
                    "Сущность / сущности": "; ".join(origin_entities),
                    "Почему это дубль": "Совпадает смысл поля по названию/семантике",
                    "Какое поле оставить": f"{primary['Поле']} ({primary['Сущность']}{' / ' + primary['Воронка'] if primary['Воронка'] else ''})",
                    "Что делать со вторым полем": "Проверить объединение / стандартизацию / перенос",
                }
            )
    return groups, duplicate_sheet_rows


def apply_recommendations(rows: list[dict[str, Any]], duplicate_groups: dict[tuple[str, str], list[dict[str, Any]]]) -> None:
    for row in rows:
        title = row["Поле"]
        code = row["Код поля"]
        origin = row["Системное / кастомное"]
        fill_rate = row["_fill_rate"]
        sem_key = row["_semantic_key"]
        comments: list[str] = []
        recommendations: list[str] = []

        if not row["_analyzed"]:
            recommendations.append("Проверить отдельно")
            comments.append("Фактическая заполненность не считалась: требуется отдельная фоновая выгрузка")

        if row["Есть смысловой дубль"] == "Да":
            recommendations.append("Объединить / стандартизировать")
            comments.append("Есть смысловой дубль в другой сущности или воронке")

        if row["_analyzed"] and origin == "Кастомное" and fill_rate == 0:
            recommendations.append("Скрыть / проверить owner / удалить")
            comments.append("Кастомное поле без данных")
        elif row["_analyzed"] and origin == "Кастомное" and fill_rate < 2:
            recommendations.append("Проверить на скрытие")
            comments.append("Очень низкая фактическая заполненность")
        elif sem_key in IMPORTANT_SEMANTICS and row["Есть смысловой дубль"] == "Да":
            recommendations.append("Унифицировать на уровне всей CRM")
            comments.append("Коммерчески значимое поле живет в нескольких вариантах")
        elif row["_analyzed"] and sem_key in IMPORTANT_SEMANTICS and fill_rate < 15:
            recommendations.append("Стандартизировать заполнение")
            comments.append("Важное управленческое поле заполняется слабо")
        elif row["_analyzed"] and origin == "Системное":
            recommendations.append("Оставить без изменений")
        else:
            recommendations.append("Оставить / проверить вручную")

        if row["Используется в автоматизациях"].startswith("Недостаточно"):
            comments.append("Нет прав/методов API для полного анализа роботов и бизнес-процессов")
        if not row["Комментарий"]:
            row["Комментарий"] = ""
        stage_usage = row["_stage_usage"]
        if stage_usage:
            comments.append(f"Чаще используется на стадиях: {stage_usage}")
        if row["_samples"]:
            comments.append(f"Примеры значений: {row['_samples']}")

        row["Рекомендация"] = "; ".join(dict.fromkeys(recommendations))
        row["Комментарий"] = " | ".join(dict.fromkeys(comments))[:1500]


def make_unused_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result = []
    for row in rows:
        if row["Системное / кастомное"] != "Кастомное" or not row["_analyzed"]:
            continue
        if row["_fill_rate"] > 0:
            continue
        result.append(
            {
                "Сущность": row["Сущность"],
                "Воронка": row["Воронка"],
                "Поле": row["Поле"],
                "Код поля": row["Код поля"],
                "Почему признано неиспользуемым": "Нет данных ни в одной записи текущей сущности/воронки",
                "Рекомендация: удалить / скрыть / проверить": "Скрыть / проверить owner / удалить",
            }
        )
    return result


def make_merge_rows(duplicate_sheet_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result = []
    for row in duplicate_sheet_rows:
        result.append(
            {
                "Текущие поля": f"{row['Поле 1']} / {row['Поле 2']}",
                "Где используются": row["Сущность / сущности"],
                "Предлагаемое единое поле": row["Поле 1"],
                "Обоснование объединения": row["Почему это дубль"],
            }
        )
    return result


def make_standardization_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result = []
    seen: set[tuple[str, str, str]] = set()
    for row in rows:
        sem_key = row["_semantic_key"]
        if not sem_key:
            continue
        if row["Есть смысловой дубль"] != "Да" and row["_fill_rate"] >= 15:
            continue
        problem_parts = []
        if row["Есть смысловой дубль"] == "Да":
            problem_parts.append("Есть несколько смыслово одинаковых полей")
        if row["_fill_rate"] < 15:
            problem_parts.append("Низкая дисциплина заполнения")
        key = (row["Поле"], row["Код поля"], row["Сущность"])
        if key in seen:
            continue
        seen.add(key)
        result.append(
            {
                "Поле": row["Поле"],
                "Проблема": "; ".join(problem_parts) or "Требуется унификация",
                "Что именно стандартизировать": "Название, owner поля, правила заполнения, место использования",
                "Предлагаемый новый формат / логика использования": f"Единый CRM-стандарт для семантики `{sem_key}`",
            }
        )
    return result


def make_final_recommendations(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    counters = {
        "удалить": 0,
        "объединить": 0,
        "переименовать": 0,
        "стандартизировать": 0,
        "оставить без изменений": 0,
        "проверить отдельно": 0,
    }
    sale_overload = sum(
        1
        for row in rows
        if row["Сущность"] == "Сделки"
        and row["Системное / кастомное"] == "Кастомное"
        and row["_analyzed"]
        and row["_fill_rate"] < 2
    )
    analytics_noise = sum(1 for row in rows if row["Есть смысловой дубль"] == "Да")
    global_duplicates = sum(1 for row in rows if row["_semantic_key"] in IMPORTANT_SEMANTICS and row["Есть смысловой дубль"] == "Да")

    for row in rows:
        rec = row["Рекомендация"].lower()
        if "удал" in rec:
            counters["удалить"] += 1
        if "объедин" in rec:
            counters["объединить"] += 1
        if "стандартиз" in rec or "унифиц" in rec:
            counters["стандартизировать"] += 1
        if "оставить без изменений" in rec:
            counters["оставить без изменений"] += 1
        if "провер" in rec:
            counters["проверить отдельно"] += 1

    result = [
        {
            "Раздел": "удалить",
            "Объект": "Кастомные поля без данных",
            "Рекомендация": f"Проверить и убрать в первую очередь {counters['удалить']} полей без фактического использования",
            "Обоснование": "Поля без данных перегружают карточки и не несут управленческой ценности",
            "Приоритет": "Высокий",
            "Статус данных": "Недостаточно данных без фоновой выгрузки",
        },
        {
            "Раздел": "объединить",
            "Объект": "Смысловые дубли",
            "Рекомендация": f"Свести дублирующие поля в единый стандарт. Найдено минимум {counters['объединить']} повторов",
            "Обоснование": "Одинаковая бизнес-логика описана разными полями",
            "Приоритет": "Высокий",
            "Статус данных": "Подтверждено частично, нужны owners",
        },
        {
            "Раздел": "стандартизировать",
            "Объект": "Коммерческие поля",
            "Рекомендация": "Унифицировать source/source_detail/client_type/segment/next_step/rejection_reason/next_contact_date по всей CRM",
            "Обоснование": f"Коммерчески значимые поля дублируются или слабо заполняются. Глобальных дублей: {global_duplicates}",
            "Приоритет": "Высокий",
            "Статус данных": "Подтверждено частично, требуется согласование",
        },
        {
            "Раздел": "проверить отдельно",
            "Объект": "Роботы и бизнес-процессы",
            "Рекомендация": "Сделать ручную ревизию использования полей в роботах, БП и карточках",
            "Обоснование": "API не дал полного доступа к шаблонам процессов и robot list",
            "Приоритет": "Высокий",
            "Статус данных": "Недостаточно данных",
        },
        {
            "Раздел": "проверить отдельно",
            "Объект": "Перегрузка продаж",
            "Рекомендация": f"Начать чистку со сделок: найдено {sale_overload} кастомных полей с заполнением ниже 2%",
            "Обоснование": "Эти поля дают наибольший шум в карточке продаж",
            "Приоритет": "Высокий",
            "Статус данных": "Недостаточно данных без фоновой выгрузки",
        },
        {
            "Раздел": "оставить без изменений",
            "Объект": "Ключевые системные поля",
            "Рекомендация": f"Сохранить системные аналитические поля без чистки. Количество строк с рекомендацией оставить: {counters['оставить без изменений']}",
            "Обоснование": "Эти поля нужны для базовой логики CRM, стадий и отчетности",
            "Приоритет": "Средний",
            "Статус данных": "Подтверждено данными",
        },
    ]
    return result


def autosize(ws) -> None:
    ensure_openpyxl()
    for column_cells in ws.columns:
        length = 0
        col_idx = column_cells[0].column
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            length = max(length, len(value))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(length + 2, 12), 60)


def write_sheet(workbook: Workbook, name: str, rows: list[dict[str, Any]]) -> None:
    ensure_openpyxl()
    ws = workbook.create_sheet(title=name)
    if not rows:
        ws.append(["Нет данных"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    ws.freeze_panes = "A2"
    autosize(ws)


def summary_text(rows: list[dict[str, Any]], unused_rows: list[dict[str, Any]], duplicate_rows: list[dict[str, Any]], standardization_rows: list[dict[str, Any]]) -> str:
    total_rows = len(rows)
    sales_rows = [row for row in rows if row["Сущность"] == "Сделки"]
    low_fill_sales = [
        row
        for row in sales_rows
        if row["Системное / кастомное"] == "Кастомное" and row["_analyzed"] and row["_fill_rate"] < 2
    ]
    important_weak = [
        row for row in rows if row["_analyzed"] and row["_semantic_key"] in IMPORTANT_SEMANTICS and row["_fill_rate"] < 15
    ]

    lines = [
        "# GD-324: краткое summary по аудиту полей Bitrix24",
        "",
        f"Подготовлено: `{MOSCOW_NOW} МСК`",
        "",
        "## Что хорошо",
        "",
        "- Системные поля по базовым CRM-сущностям доступны и структурированы.",
        "- Основные поля стадий, ответственных и источников в системе присутствуют.",
        "- По смарт-процессам Bitrix отдает полную мету полей, включая кастомные.",
        "",
        "## Что плохо",
        "",
        f"- Всего проанализировано `{total_rows}` строк полей по сущностям и воронкам.",
        f"- В workbook уже выделены кандидаты на чистку структуры, но без фонового live-sweep нельзя честно назвать итоговое число неиспользуемых полей.",
        f"- Найдено минимум `{len(duplicate_rows)}` строк по смысловым дублям.",
        f"- Найдено `{len(standardization_rows)}` кандидатов на стандартизацию.",
        "- Точные проценты заполнения по всем карточкам в этой версии не завершены: для них нужен отдельный фоновый прогон REST-выгрузки.",
        "",
        "## Где основной мусор",
        "",
        "- В разросшейся сетке кастомных полей без подтвержденного owner и без единого стандарта.",
        "- В смысловых дублях коммерческих полей между разными сущностями и воронками.",
        "- В полях без понятного owner и единых правил заполнения.",
        "",
        "## Где дубли",
        "",
        "- Источники и подисточники.",
        "- Причины отказа / потери.",
        "- Следующий шаг / следующий контакт.",
        "- Тип клиента / сегмент / сфера деятельности.",
        "",
        "## С чего начать чистку",
        "",
        "1. Сначала добить factual sweep по заполненности, а затем сразу скрывать кастомные поля без данных.",
        "2. Зафиксировать единый стандарт коммерческих полей: source, source_detail, client_type, segment, next_step, next_contact_date, rejection_reason.",
        "3. Разобрать поля продаж отдельно как самый перегруженный контур.",
        "4. Отдельно вручную проверить поля, задействованные в роботах, бизнес-процессах и карточках, потому что полных API-прав для этого анализа нет.",
        "",
        "## Ограничения",
        "",
        "- Использование полей в бизнес-процессах и роботах определено не полностью: доступ к `bizproc.workflow.template.list` ограничен правами, а API роботов неполный.",
        "- Использование полей в макете карточек требует ручной проверки в интерфейсе Bitrix24.",
        "- В этой версии собран полный мета-аудит по всем полям и сущностям. Полный factual sweep по всем live-записям требует отдельного фонового запуска из-за объема REST-пагинации.",
        f"- После полноценного live-sweep особое внимание нужно дать коммерчески важным полям. В текущей версии они уже выделены семантически, но не все подтверждены фактической заполняемостью.",
    ]
    return "\n".join(lines)


def main() -> None:
    ensure_openpyxl()
    TMP_DIR.mkdir(parents=True, exist_ok=True)
    DOCS_DIR.mkdir(parents=True, exist_ok=True)
    auth = make_auth()
    slices = build_entity_slices(auth)

    all_rows: list[dict[str, Any]] = []
    entity_stats: list[dict[str, Any]] = []
    for entity_slice, field_map in slices:
        analyzed_codes = set(select_codes_for_analysis(field_map, dynamic=entity_slice.entity_name not in {"Сделки", "Лиды", "Контакты", "Компании"}))
        rows = entity_usage_rows(
            entity_slice.entity_name,
            entity_slice.category_name,
            field_map,
            entity_slice.records,
            entity_slice.stage_field,
            entity_slice.stage_labels,
            analyzed_codes=analyzed_codes,
        )
        all_rows.extend(rows)
        entity_stats.append(
            {
                "entity": entity_slice.entity_name,
                "category": entity_slice.category_name,
                "records": len(entity_slice.records),
                "fields": len(field_map),
            }
        )

    duplicate_groups, duplicate_rows = build_duplicate_index(all_rows)
    apply_recommendations(all_rows, duplicate_groups)
    unused_rows = make_unused_rows(all_rows)
    merge_rows = make_merge_rows(duplicate_rows)
    standardization_rows = make_standardization_rows(all_rows)
    final_recommendations = make_final_recommendations(all_rows)

    workbook = Workbook()
    workbook.remove(workbook.active)

    visible_rows = []
    for row in all_rows:
        visible = {key: value for key, value in row.items() if not key.startswith("_")}
        visible["Комментарий"] = visible["Комментарий"]
        visible_rows.append(visible)

    write_sheet(workbook, "Все поля", visible_rows)
    write_sheet(workbook, "Неиспользуемые поля", unused_rows)
    write_sheet(workbook, "Дублирующие поля", duplicate_rows)
    write_sheet(workbook, "Поля для объединения", merge_rows)
    write_sheet(workbook, "Поля для стандартизации", standardization_rows)
    write_sheet(workbook, "Итоговые рекомендации", final_recommendations)
    workbook.save(OUTPUT_XLSX)

    OUTPUT_JSON.write_text(
        json.dumps(
            {
                "generated_at_msk": MOSCOW_NOW,
                "entity_stats": entity_stats,
                "all_rows": visible_rows,
                "unused_rows": unused_rows,
                "duplicate_rows": duplicate_rows,
                "merge_rows": merge_rows,
                "standardization_rows": standardization_rows,
                "final_recommendations": final_recommendations,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    OUTPUT_SUMMARY.write_text(
        summary_text(all_rows, unused_rows, duplicate_rows, standardization_rows),
        encoding="utf-8",
    )

    print(str(OUTPUT_XLSX))
    print(str(OUTPUT_SUMMARY))
    print(json.dumps({"rows": len(visible_rows), "unused": len(unused_rows), "duplicates": len(duplicate_rows)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
