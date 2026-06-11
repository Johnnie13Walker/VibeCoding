#!/usr/bin/env python3
"""Генератор клиентской сметы — главный выигрышный формат КП (WON-KP-INSIGHTS.md).

Постатейная смета «работа → часы → деньги» в .xlsx: ТОЛЬКО клиентский лист
(антипаттерн утечки внутренних калькуляторов закрыт по построению), цены без НДС
→ +НДС 5% (решение 10.06: все сделки без указания юрлица) → каскад скидок
(10% логотип при разработке, 5% оперативная оплата с дедлайном). Контроль
минимального чека: скидку на минчек не применяем (правило РОП 21.05.2026).

    python3 kp_smeta.py --init clients/<клиент> [--service seo|program|orm|ppc]
    python3 kp_smeta.py clients/<клиент>        # smeta.json → Смета_<клиент>.xlsx

Тарифы сняты с «Продуктовой Матрицы 2026» (Google Sheet 1wyNlRjV…) 10.06.2026 —
перед отправкой клиенту сверить с актуальной таблицей.
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import date
from pathlib import Path

VAT = 0.05            # решение 10.06: НДС 5% без указания юрлица
DISC_LOGO = 0.10      # за размещение логотипа агентства — только при разработке сайта
DISC_FAST = 0.05      # за оплату в течение 10 календарных дней

# минимальные чеки направлений (матрица, снято 10.06.2026)
MIN_CHECK = {
    "seo": 75_000,           # визитка/лендинг; витрина 80, услуги 85, магазин 115, портал 135
    "program": 31_000,
    "orm": 60_000,
    "ppc": 62_500,
    "wd": 900_000, "wdt": 450_000, "tv": 480_000, "ta": 720_000,
    "lp": 140_000, "branding": 77_500,
}

# заготовки строк по услугам (тарифы матрицы, сняты 10.06.2026; состав правит сейлс)
# строка-«секция» {"section": "..."} рисует заголовок этапа; итог этапа считается сам
PRESETS = {
    "seo": {"service": "seo", "items": [
        {"name": "SEO-продвижение: техника, контент, ссылки (42 ч/мес)", "monthly": 85_000},
        {"name": "GEO: карточки Яндекс.Карт и геосервисов", "monthly": 0},
    ]},
    # типовые работы 1-го месяца — из вопросов брифа техподдержки СП1056; списываются
    # из пакета часов тарифа, поэтому идут строками «включено» без отдельной цены
    "program": {"service": "program", "items": [
        {"name": "Техподдержка, тариф «Цикл» (10 ч/мес)", "hours": 10, "rate": 3_100},
        {"section": "ТИПОВЫЕ РАБОТЫ ПЕРВОГО МЕСЯЦА (в рамках пакета часов)"},
        {"name": "Диагностика сайта и приоритизация задач", "included": True},
        {"name": "Обновление CMS и модулей, настройка резервных копий", "included": True},
        {"name": "Поиск и устранение ошибок вёрстки и логики", "included": True},
        {"name": "Оптимизация скорости загрузки", "included": True},
        {"name": "Мелкие правки контента и баннеров", "included": True},
        {"name": "Взаимодействие с хостинг-провайдером", "included": True},
    ]},
    "program-deposit": {"service": "program", "items": [
        {"name": "Техподдержка, тариф «Депозит» (10 ч/мес)", "hours": 10, "rate": 3_900},
        {"section": "ТИПОВЫЕ РАБОТЫ ПЕРВОГО МЕСЯЦА (в рамках пакета часов)"},
        {"name": "Диагностика сайта и приоритизация задач", "included": True},
        {"name": "Обновление CMS и модулей, настройка резервных копий", "included": True},
        {"name": "Поиск и устранение ошибок вёрстки и логики", "included": True},
        {"name": "Оптимизация скорости загрузки", "included": True},
        {"name": "Мелкие правки контента и баннеров", "included": True},
        {"name": "Взаимодействие с хостинг-провайдером", "included": True},
    ]},
    "orm": {"service": "orm", "items": [
        {"name": "Управление репутацией, тариф «Старт» (от 3 карточек)", "monthly": 60_000},
    ]},
    "ppc": {"service": "ppc", "items": [
        {"name": "Ведение контекстной рекламы (фикс, бюджет 150–300 тыс)", "monthly": 65_000},
        {"name": "Рекламный бюджет Яндекса — оплачивается напрямую, в смету не входит", "monthly": 0},
    ]},
    # сайт на шаблоне 1С-Битрикс (вкладка «ТВ»; ставка 2 800 ₽/ч; итог матрицы 681 100)
    "tv": {"service": "tv", "items": [
        {"section": "ЭТАП 1: ДИЗАЙН"},
        {"name": "Разработка технического задания", "hours": 12, "rate": 2_800},
        {"name": "Покупка шаблона сайта (medberry.website)", "once": 69_000},
        {"name": "Разворот шаблона на хостинге, настройка под 1С-Битрикс", "hours": 12, "rate": 2_800},
        {"name": "Подбор и покупка фотоконтента (5 шт)", "once": 14_000},
        {"name": "Отрисовка баннеров под акции (до 5 шт)", "hours": 10, "rate": 2_800},
        {"name": "Корректировка цветовой гаммы шаблона", "hours": 8, "rate": 2_800},
        {"name": "Корректировка страницы «Услуга»", "hours": 10, "rate": 2_800},
        {"name": "Корректировка главной страницы", "hours": 14, "rate": 2_800},
        {"section": "ЭТАП 2: ВЁРСТКА И ПРОГРАММИРОВАНИЕ"},
        {"name": "Вёрстка индивидуальных компонентов (десктоп + адаптив)", "hours": 20, "rate": 2_800},
        {"name": "Программная связка всех элементов сайта", "hours": 22, "rate": 2_800},
        {"name": "Лицензия 1С-Битрикс «Стандарт»", "once": 20_500},
        {"name": "Версия для слабовидящих", "hours": 30, "rate": 2_800},
        {"section": "ЭТАП 3: НАПОЛНЕНИЕ КОНТЕНТОМ"},
        {"name": "Подготовка ТЗ для статей (4 шт)", "once": 5_200},
        {"name": "Тексты услуговых страниц (4 шт)", "once": 20_800},
        {"name": "Наполнение контентом", "hours": 30, "rate": 2_800},
        {"name": "Перенос сайта на рабочий домен", "hours": 4, "rate": 2_800},
        {"name": "Менеджмент проекта", "hours": 10, "rate": 2_800},
        {"name": "Тестирование сайта", "hours": 15, "rate": 2_800},
    ]},
    # интернет-магазин «под ключ» на 1С-Битрикс (вкладка «TA»; итог матрицы 676 101)
    "ta": {"service": "ta", "items": [
        {"section": "ЭТАП 1: ДИЗАЙН"},
        {"name": "Разработка технического задания", "hours": 12, "rate": 2_800},
        {"name": "Покупка шаблона сайта (acoola-shop.ru, спецпредложение)", "once": 1},
        {"name": "Разворот шаблона на хостинге, настройка под 1С-Битрикс", "hours": 12, "rate": 2_800},
        {"name": "Подбор и покупка фотоконтента (5 шт)", "once": 14_000},
        {"name": "Отрисовка баннеров под спецпредложения (до 5 шт)", "hours": 10, "rate": 2_800},
        {"name": "Корректировка цветовой гаммы шаблона", "hours": 14, "rate": 2_800},
        {"name": "Корректировка страницы «Карточка товара»", "hours": 8, "rate": 2_800},
        {"name": "Корректировка главной страницы", "hours": 14, "rate": 2_800},
        {"section": "ЭТАП 2: ВЁРСТКА И ПРОГРАММИРОВАНИЕ (VUE)"},
        {"name": "Вёрстка индивидуальных компонентов (десктоп + адаптив)", "hours": 20, "rate": 2_800},
        {"name": "Программная связка всех элементов сайта", "hours": 35, "rate": 2_800},
        {"name": "Лицензия 1С-Битрикс «Малый бизнес»", "once": 47_000},
        {"section": "ЭТАП 3: НАПОЛНЕНИЕ КОНТЕНТОМ"},
        {"name": "Подготовка ТЗ для статей (4 шт)", "once": 5_200},
        {"name": "Тексты услуговых страниц (4 шт)", "once": 20_800},
        {"name": "Перенос контента с текущего сайта", "hours": 16, "rate": 2_800},
        {"name": "Интеграция системы оплаты (банк)", "once": 46_000},
        {"name": "Интеграция системы оплаты (СБП)", "once": 16_100},
        {"name": "Интеграция системы доставки", "once": 23_000},
        {"name": "Перенос сайта на рабочий домен", "hours": 4, "rate": 2_800},
        {"name": "Менеджмент проекта", "hours": 20, "rate": 2_800},
        {"name": "Тестирование сайта", "hours": 15, "rate": 2_800},
    ]},
    # посадочная страница (вкладка «LP»; итог матрицы 145 300)
    "lp": {"service": "lp", "items": [
        {"name": "Разработка дизайна", "hours": 20, "rate": 2_800},
        {"name": "Вёрстка посадочной страницы (десктоп + адаптив)", "hours": 15, "rate": 2_800},
        {"name": "Наполнение контентом", "hours": 4, "rate": 2_800},
        {"name": "Написание контента (до 5 000 символов)", "once": 10_000},
        {"name": "Корректировка и редактирование текста", "hours": 3, "rate": 2_500},
        {"name": "Тестирование и проверка форм записи", "hours": 2, "rate": 3_100},
        {"name": "Перенос на рабочий хостинг и домен", "hours": 4, "rate": 3_100},
    ]},
    # фирменный стиль (вкладка «Branding»; ставка 3 100 ₽/ч; итог матрицы 517 700)
    "branding": {"service": "branding", "items": [
        {"name": "Логотип (3 первоначальных варианта, 3 доработки одного)", "hours": 25, "rate": 3_100},
        {"name": "Подбор цветов и шрифтов", "hours": 10, "rate": 3_100},
        {"name": "Фирменный бланк", "hours": 8, "rate": 3_100},
        {"name": "Шаблон презентации", "hours": 10, "rate": 3_100},
        {"name": "Визитка", "hours": 10, "rate": 3_100},
        {"name": "Наружная вывеска", "hours": 14, "rate": 3_100},
        {"name": "Упаковка товара", "hours": 20, "rate": 3_100},
        {"name": "Гайдлайн", "hours": 70, "rate": 3_100},
    ]},
}


# ── расчёты (pure, под тестами) ──────────────────────────────────────────────

def line_total(item: dict) -> float:
    """Стоимость строки: часы×ставка либо фикс (monthly/once)."""
    if item.get("hours") is not None and item.get("rate") is not None:
        return float(item["hours"]) * float(item["rate"])
    return float(item.get("monthly") or item.get("once") or 0)


def cascade(subtotal: float, logo: bool = False, fast: bool = False) -> list[tuple[str, float]]:
    """Каскад итогов в рублях: без НДС → с НДС → скидки по очереди.

    Подача из выигравших смет: каждый шаг каскада — отдельная строка с суммой.
    """
    steps = [("Итого без НДС", round(subtotal, 2))]
    cur = subtotal * (1 + VAT)
    steps.append((f"Итого с НДС {int(VAT*100)}%", round(cur, 2)))
    if logo:
        cur *= 1 - DISC_LOGO
        steps.append((f"− {int(DISC_LOGO*100)}% за размещение логотипа агентства", round(cur, 2)))
    if fast:
        cur *= 1 - DISC_FAST
        steps.append((f"− {int(DISC_FAST*100)}% за оплату в течение 10 дней", round(cur, 2)))
    return steps


def min_check_guard(subtotal: float, service: str | None,
                    discounts_requested: bool) -> str | None:
    """Правила минимального чека: ниже минчека не продаём; скидка на минчек запрещена."""
    mc = MIN_CHECK.get(service or "")
    if not mc:
        return None
    if subtotal < mc:
        return (f"⛔ Сумма {subtotal:,.0f} ₽ ниже минимального чека направления "
                f"{mc:,.0f} ₽ — по правилу матрицы сделку не удешевляем")
    if discounts_requested and subtotal == mc:
        return "⚠ Скидка на минимальный чек запрещена (правило матрицы) — скидки не применены"
    return None


def build_rows(items: list[dict]) -> tuple[list[dict], float]:
    """Строки к отрисовке (item / section / section_total) + общий итог без НДС.

    Итог этапа добавляется автоматически перед следующей секцией и в конце,
    как в выигравших сметах («ИТОГО ЗА ЭТАП …»).
    """
    rows, subtotal = [], 0.0
    sec_name, sec_sum = None, 0.0

    def close_section():
        nonlocal sec_name, sec_sum
        # секция из «включено»-строк денег не несёт — нулевой подытог не печатаем
        if sec_name is not None and sec_sum > 0:
            rows.append({"kind": "section_total",
                         "name": f"Итого за {sec_name.lower()}", "total": round(sec_sum, 2)})
        sec_name, sec_sum = None, 0.0

    for it in items:
        if "section" in it:
            close_section()
            sec_name, sec_sum = it["section"], 0.0
            rows.append({"kind": "section", "name": it["section"]})
            continue
        total = line_total(it)
        subtotal += total
        sec_sum += total
        rows.append({"kind": "item", "item": it, "total": round(total, 2)})
    close_section()
    return rows, round(subtotal, 2)


def validate_smeta(spec: dict) -> list[str]:
    """Ошибки спеки до генерации: пустые строки, отрицательные суммы, нет клиента."""
    errs = []
    if not spec.get("client"):
        errs.append("нет поля client")
    items = spec.get("items") or []
    if not items:
        errs.append("нет строк items")
    for i, it in enumerate(items, 1):
        if "section" in it:
            if not it["section"]:
                errs.append(f"строка {i}: пустое имя секции")
            continue
        if not it.get("name"):
            errs.append(f"строка {i}: нет name")
        if line_total(it) < 0:
            errs.append(f"строка {i}: отрицательная сумма")
    return errs


# ── xlsx ─────────────────────────────────────────────────────────────────────

BRAND_STYLE = {
    "Belberry": {"accent": "5B50D6", "name": "Belberry — маркетинг для медицины"},
    "Acoola Team": {"accent": "3086FB", "name": "Acoola Team — digital-агентство"},
}


def write_xlsx(spec: dict, out_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    brand = spec.get("brand") or "Belberry"
    st = BRAND_STYLE.get(brand, BRAND_STYLE["Belberry"])
    accent, ink, line = st["accent"], "1D1D1F", "ECEAE3"

    wb = Workbook()
    ws = wb.active
    ws.title = "Смета"
    widths = [52, 10, 12, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    thin = Side(style="thin", color=line)
    border = Border(bottom=thin)
    money = '#,##0 "₽"'

    ws["A1"] = f"Смета — {spec['client']}"
    ws["A1"].font = Font(size=16, bold=True, color=ink)
    ws["A2"] = st["name"]
    ws["A2"].font = Font(size=10, color="6E6E73")
    ws["A3"] = f"Дата: {spec.get('date') or date.today().strftime('%d.%m.%Y')}"
    ws["A3"].font = Font(size=10, color="6E6E73")

    r = 5
    for col, head in enumerate(["Работы", "Часы", "Ставка", "Стоимость, без НДС"], 1):
        c = ws.cell(row=r, column=col, value=head)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=accent)
        c.alignment = Alignment(horizontal="left" if col == 1 else "right")
    r += 1

    rows, subtotal = build_rows(spec["items"])
    soft = PatternFill("solid", fgColor="F2F2F4")
    for row in rows:
        if row["kind"] == "section":
            c = ws.cell(row=r, column=1, value=row["name"])
            c.font = Font(bold=True, size=10, color=accent)
            for col in range(1, 5):
                ws.cell(row=r, column=col).fill = soft
            r += 1
            continue
        if row["kind"] == "section_total":
            lc = ws.cell(row=r, column=1, value=row["name"])
            vc = ws.cell(row=r, column=4, value=row["total"])
            vc.number_format = money
            vc.alignment = Alignment(horizontal="right")
            lc.font = vc.font = Font(bold=True, size=10, color=ink)
            for col in range(1, 5):
                ws.cell(row=r, column=col).border = border
            r += 1
            continue
        it = row["item"]
        ws.cell(row=r, column=1, value=it["name"]).alignment = Alignment(wrap_text=True)
        if it.get("hours") is not None:
            ws.cell(row=r, column=2, value=it["hours"]).alignment = Alignment(horizontal="right")
            rc = ws.cell(row=r, column=3, value=it.get("rate"))
            rc.number_format = money
            rc.alignment = Alignment(horizontal="right")
        if it.get("included"):
            tc = ws.cell(row=r, column=4, value="включено")
            tc.font = Font(size=9, color="6E6E73")
        else:
            tc = ws.cell(row=r, column=4, value=row["total"])
            tc.number_format = money
        tc.alignment = Alignment(horizontal="right")
        for col in range(1, 5):
            ws.cell(row=r, column=col).border = border
        r += 1

    flags = spec.get("flags") or {}
    guard = min_check_guard(subtotal, spec.get("service"),
                            bool(flags.get("logo_discount") or flags.get("fast_pay")))
    apply_disc = not guard
    r += 1
    for label, value in cascade(subtotal,
                                logo=apply_disc and bool(flags.get("logo_discount")),
                                fast=apply_disc and bool(flags.get("fast_pay"))):
        lc = ws.cell(row=r, column=1, value=label)
        vc = ws.cell(row=r, column=4, value=value)
        vc.number_format = money
        vc.alignment = Alignment(horizontal="right")
        lc.font = vc.font = Font(bold=True, color=ink)
        r += 1

    r += 1
    notes = ["Цены приведены без учёта рекламных бюджетов площадок.",
             "Скидка за оперативную оплату действует 10 календарных дней с даты сметы."]
    if spec.get("deadline"):
        notes.append(f"Срок действия предложения: до {spec['deadline']}.")
    if guard:
        notes.append(guard.lstrip("⛔⚠ "))
    for n in notes:
        ws.cell(row=r, column=1, value=n).font = Font(size=9, color="6E6E73")
        r += 1

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    if guard:
        print(guard)
    print(f"✅ {out_path}")


# ── CLI ──────────────────────────────────────────────────────────────────────

def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("client_dir", type=Path, help="папка клиента (clients/<имя>)")
    p.add_argument("--init", action="store_true", help="создать заготовку smeta.json")
    p.add_argument("--service", choices=sorted(PRESETS), default="seo")
    a = p.parse_args()

    spec_path = a.client_dir / "smeta.json"
    if a.init:
        if spec_path.exists():
            sys.exit(f"⚠ {spec_path} уже есть — правь её, не перезаписываю")
        spec = {"client": a.client_dir.name, "brand": "Belberry",
                "service": a.service, "deadline": "",
                "flags": {"logo_discount": False, "fast_pay": True},
                **PRESETS[a.service]}
        a.client_dir.mkdir(parents=True, exist_ok=True)
        spec_path.write_text(json.dumps(spec, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"✅ заготовка: {spec_path} — впиши состав и суммы из калькулятора, затем "
              f"запусти без --init")
        return 0

    if not spec_path.exists():
        sys.exit(f"нет {spec_path} — сначала: kp_smeta.py --init {a.client_dir}")
    spec = json.loads(spec_path.read_text(encoding="utf-8"))
    errs = validate_smeta(spec)
    if errs:
        sys.exit("ошибки smeta.json: " + "; ".join(errs))
    write_xlsx(spec, a.client_dir / f"Смета_{spec['client']}.xlsx")
    return 0


if __name__ == "__main__":
    sys.exit(main())
