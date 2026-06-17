#!/usr/bin/env python3
from __future__ import annotations

import json
from collections import Counter, defaultdict
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


TEMPLATE = Path("/Users/pro2kuror/Downloads/Клиенты_визуал.xlsx")
DB_JSON = Path("/tmp/portfolio_db_values.json")
CLASS_JSON = Path("/tmp/portfolio_classification_values.json")
OUTPUT = Path("/tmp/Портфолио Belberry и Acoola Team.xlsx")

BRANDS = {"Belberry", "Acoola Team"}
ACTIVE_YEARS = {2025, 2026}
HEADER_FILL = "17324D"
BLUE = "0F3B66"
GRID = "D9E2EF"
LIGHT = "F6F8FB"
WHITE = "FFFFFF"


def safe(value) -> str:
    return str(value or "").strip()


def to_int(value) -> int:
    try:
        return int(float(str(value).replace(",", ".").strip()))
    except Exception:
        return 0


def period(years: set[int]) -> str:
    values = sorted(years)
    if not values:
        return ""
    if values[0] == values[-1]:
        return str(values[0])
    return f"{values[0]}-{values[-1]}"


def years_text(years: set[int]) -> str:
    return ", ".join(str(y) for y in sorted(years))


def status(max_year: int) -> str:
    if max_year >= 2025:
        return "Активен"
    if max_year == 2024:
        return "Спящий"
    return "Архив"


def pct(value: float, total: float) -> float:
    return value / total if total else 0


def top(counter: Counter) -> tuple[str, int]:
    if not counter:
        return "", 0
    return sorted(counter.items(), key=lambda item: (-item[1], item[0]))[0]


def load_values(path: Path, key: str) -> list[list]:
    payload = json.loads(path.read_text())
    return payload[key]


def load_model():
    rows = load_values(DB_JSON, "Данные!A:H")
    class_rows = load_values(CLASS_JSON, "Классификация!A:F")
    classification = {}
    for row in class_rows[1:]:
        project = safe(row[0] if len(row) > 0 else "")
        if not project:
            continue
        classification[project] = {
            "category": safe(row[1] if len(row) > 1 else "") or "Не определено",
            "subcategory": safe(row[2] if len(row) > 2 else "") or "Не определено",
            "source": safe(row[3] if len(row) > 3 else "") or "сохранено",
            "source_rows": safe(row[4] if len(row) > 4 else ""),
            "source_row_numbers": safe(row[5] if len(row) > 5 else ""),
        }

    projects = {}
    years = set()
    source_rows_by_project = defaultdict(list)
    for index, row in enumerate(rows[1:], start=2):
        project = safe(row[0] if len(row) > 0 else "")
        service = safe(row[1] if len(row) > 1 else "")
        brand = safe(row[2] if len(row) > 2 else "")
        year = to_int(row[3] if len(row) > 3 else "")
        month = to_int(row[4] if len(row) > 4 else "")
        department = safe(row[6] if len(row) > 6 else "")
        new_old = safe(row[7] if len(row) > 7 else "")
        if not project or not service or not year or brand not in BRANDS:
            continue
        years.add(year)
        source_rows_by_project[project].append(str(index))
        if project not in projects:
            cls = classification.get(project, {})
            projects[project] = {
                "project": project,
                "category": cls.get("category", "Не определено"),
                "subcategory": cls.get("subcategory", "Не определено"),
                "class_source": cls.get("source", "нет классификации"),
                "class_rows": cls.get("source_rows", ""),
                "class_row_numbers": cls.get("source_row_numbers", ""),
                "brands": set(),
                "departments": set(),
                "new_old": set(),
                "services": set(),
                "years": set(),
                "months": set(),
                "receipts": 0,
                "services_by_year": defaultdict(set),
                "years_by_service": defaultdict(set),
                "months_by_service": defaultdict(set),
                "receipts_by_service": Counter(),
            }
        item = projects[project]
        item["brands"].add(brand)
        item["departments"].add(department)
        item["new_old"].add(new_old)
        item["services"].add(service)
        item["years"].add(year)
        if month:
            item["months"].add(f"{year}-{month:02d}")
            item["months_by_service"][service].add(f"{year}-{month:02d}")
        item["services_by_year"][year].add(service)
        item["years_by_service"][service].add(year)
        item["receipts_by_service"][service] += 1
        item["receipts"] += 1

    for project, item in projects.items():
        item["min_year"] = min(item["years"])
        item["max_year"] = max(item["years"])
        item["period"] = period(item["years"])
        item["status"] = status(item["max_year"])
        item["month_count"] = len(item["months"])
        item["services_list"] = sorted(item["services"])
        if item["class_source"] == "нет классификации":
            item["class_rows"] = str(len(source_rows_by_project[project]))
            item["class_row_numbers"] = ",".join(source_rows_by_project[project][:40])

    return {
        "projects": sorted(projects.values(), key=lambda x: (x["category"], x["project"])),
        "years": sorted(years),
    }


def copy_row_style(ws, src_row: int, dst_row: int, max_col: int):
    for col in range(1, max_col + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format


def clear_sheet_values(ws, start_row: int, max_row: int, max_col: int):
    for row in ws.iter_rows(min_row=start_row, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def write_matrix(ws, start_row: int, start_col: int, values: list[list], style_row: int | None = None):
    max_col = max((len(row) for row in values), default=0)
    for r_index, row in enumerate(values, start=start_row):
        if style_row:
            copy_row_style(ws, style_row, r_index, max_col)
        for c_index, value in enumerate(row, start=start_col):
            ws.cell(r_index, c_index).value = value


def recreate_table(ws, name: str, ref: str):
    for table_name in list(ws.tables):
        del ws.tables[table_name]
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def style_header(ws, row: int, cols: int):
    fill = PatternFill("solid", fgColor=HEADER_FILL)
    font = Font(color=WHITE, bold=True)
    border = Border(bottom=Side(style="thin", color=GRID))
    for col in range(1, cols + 1):
        cell = ws.cell(row, col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def rebuild_charts(wb):
    ws_dash = wb["Дашборд"]
    ws_cat = wb["Категории"]
    ws_prod = wb["Продукты"]
    ws_data = wb["dash_data"]
    ws_dash._charts = []
    ws_cat._charts = []
    ws_prod._charts = []

    line = LineChart()
    line.title = "Динамика портфолио по годам"
    line.style = 13
    line.height = 9
    line.width = 17
    line.y_axis.title = "Проектов"
    line.x_axis.title = "Год"
    data = Reference(ws_data, min_col=2, max_col=7, min_row=1, max_row=12)
    cats = Reference(ws_data, min_col=1, min_row=2, max_row=12)
    line.add_data(data, titles_from_data=True)
    line.set_categories(cats)
    ws_dash.add_chart(line, "J17")

    bar = BarChart()
    bar.title = "Проектов по категориям"
    bar.style = 10
    bar.height = 12
    bar.width = 15
    bar.type = "bar"
    bar.y_axis.title = "Категория"
    bar.x_axis.title = "Проектов"
    cat_rows = min(ws_cat.max_row, 16)
    data = Reference(ws_cat, min_col=2, min_row=4, max_row=cat_rows)
    cats = Reference(ws_cat, min_col=1, min_row=5, max_row=cat_rows)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    ws_cat.add_chart(bar, "J4")

    product_line = LineChart()
    product_line.title = "Динамика топ-продуктов"
    product_line.style = 13
    product_line.height = 10
    product_line.width = 17
    data = Reference(ws_data, min_col=2, max_col=7, min_row=1, max_row=12)
    cats = Reference(ws_data, min_col=1, min_row=2, max_row=12)
    product_line.add_data(data, titles_from_data=True)
    product_line.set_categories(cats)
    ws_prod.add_chart(product_line, "A30")


def build_rows(model):
    projects = model["projects"]
    years = model["years"]
    categories = []
    category_map = defaultdict(list)
    for item in projects:
        category_map[item["category"]].append(item)
    for category, items in category_map.items():
        sub = Counter(item["subcategory"] for item in items)
        products = Counter(service for item in items for service in item["services"])
        active = sum(1 for item in items if item["years"] & ACTIVE_YEARS)
        categories.append([
            category,
            len(items),
            pct(len(items), len(projects)),
            len(sub),
            top(sub)[0],
            top(products)[0],
            active,
            sum(item["month_count"] for item in items) / len(items),
        ])
    categories.sort(key=lambda row: (-row[1], row[0]))

    services = sorted({service for item in projects for service in item["services"]})
    products = []
    for service in services:
        service_projects = [item for item in projects if service in item["services"]]
        cat = Counter(item["category"] for item in service_projects)
        by_year = Counter()
        for item in service_projects:
            for year in item["years_by_service"][service]:
                by_year[year] += 1
        top_cat, top_cat_count = top(cat)
        delta = by_year[2025] - by_year[2024]
        products.append([
            service,
            len(service_projects),
            pct(len(service_projects), len(projects)),
            f"{top_cat} ({top_cat_count})" if top_cat else "",
            *[by_year[year] for year in years],
            delta,
        ])
    products.sort(key=lambda row: (-row[1], row[0]))

    clients = []
    for item in projects:
        clients.append([
            item["project"],
            item["category"],
            item["subcategory"],
            item["period"],
            ", ".join(item["services_list"]),
            *[", ".join(sorted(item["services_by_year"][year])) for year in years],
            item["month_count"],
            item["receipts"],
            item["status"],
        ])

    product_client = []
    for item in projects:
        for service in item["services_list"]:
            product_client.append([
                service,
                item["project"],
                item["category"],
                item["subcategory"],
                years_text(item["years_by_service"][service]),
                period(item["years_by_service"][service]),
                years_text(item["years"]),
                ", ".join(item["services_list"]),
                len(item["months_by_service"][service]),
                item["receipts_by_service"][service],
                "нет классификации" if item["class_source"] == "нет классификации" else "есть данные",
            ])
    product_client.sort(key=lambda row: (row[0], row[2], row[1]))

    classification = [[
        item["project"],
        item["category"],
        item["subcategory"],
        item["class_source"],
        item["class_rows"],
        item["class_row_numbers"],
    ] for item in projects]

    retention = []
    cohorts = defaultdict(list)
    for item in projects:
        cohorts[item["min_year"]].append(item)
    for cohort in sorted(cohorts):
        items = cohorts[cohort]
        retention.append([
            cohort,
            len(items),
            *[(sum(1 for item in items if year in item["years"]) / len(items) if year >= cohort else None) for year in years],
        ])

    top_products = [row[0] for row in products[:5]]
    dynamics = []
    for year in years:
        counts = Counter()
        for item in projects:
            for service in item["services_by_year"][year]:
                counts[service] += 1
        top_sum = sum(counts[p] for p in top_products)
        dynamics.append([year, *[counts[p] for p in top_products], sum(counts.values()) - top_sum])

    heatmap_products = [row[0] for row in products[:10]]
    heatmap = []
    for category, items in category_map.items():
        heatmap.append([
            category,
            len(items),
            *[sum(1 for item in items if product in item["services"]) for product in heatmap_products],
        ])
    heatmap.sort(key=lambda row: (-row[1], row[0]))

    return {
        "categories": categories,
        "products": products,
        "clients": clients,
        "product_client": product_client,
        "classification": classification,
        "retention": retention,
        "dynamics": dynamics,
        "heatmap": heatmap,
        "years": years,
        "top_products": top_products,
    }


def update_workbook():
    model = load_model()
    built = build_rows(model)
    wb = load_workbook(TEMPLATE)

    # Дашборд
    ws = wb["Дашборд"]
    clear_sheet_values(ws, 1, 90, 17)
    top_category = built["categories"][0]
    top_product = built["products"][0]
    active = sum(1 for item in model["projects"] if item["years"] & ACTIVE_YEARS)
    years = built["years"]
    ws["B2"] = "Портфолио Belberry и Acoola Team"
    ws["B3"] = f"{len(model['projects'])} клиентов · {len(built['categories'])} категорий · {len(built['products'])} продуктов · {years[0]}–{years[-1]}"
    ws["B5"] = "ГЛАВНОЕ"
    ws["B6"] = f"•  Топ-ниша: «{top_category[0]}» ({top_category[1]} клиентов, {top_category[2]:.0%} всего). Внутри: {top_category[4]}."
    ws["B7"] = f"•  Топ-продукт: {top_product[0]} ({top_product[1]} проектов, {top_product[2]:.0%} портфолио)."
    ws["B8"] = f"•  Активных в 2025–26: {active} клиентов ({active / len(model['projects']):.0%}). Источник: закрытая база без сумм оплат."
    kpis = [
        ["КЛИЕНТОВ", "АКТИВНЫХ В 2025–26", "ТОП-НИША", "ТОП-ПРОДУКТ"],
        [len(model["projects"]), f"{active} ({active / len(model['projects']):.0%})", top_category[0], top_product[0]],
        ["всего в портфолио", "сейчас в работе", f"{top_category[2]:.0%} · {top_category[1]} клиентов", f"{top_product[1]} проектов · {top_product[2]:.0%}"],
    ]
    for row_idx, row in enumerate(kpis, start=11):
        for start_col, value in zip([2, 6, 10, 14], row):
            ws.cell(row_idx, start_col).value = value

    ws["B16"] = "УДЕРЖАНИЕ КЛИЕНТОВ ПО КОГОРТАМ"
    retention_table = [["Когорта", "Размер", *years], *built["retention"]]
    write_matrix(ws, 17, 2, retention_table, style_row=18)
    for row in ws.iter_rows(min_row=18, max_row=17 + len(retention_table), min_col=4, max_col=3 + len(years)):
        for cell in row:
            if cell.value is not None:
                cell.number_format = "0%"
    ws["B31"] = "Динамика по топ-продуктам"
    write_matrix(ws, 32, 2, [["Год", *built["top_products"], "Другие"], *built["dynamics"]], style_row=33)
    ws["B45"] = "ТЕПЛОВАЯ КАРТА: КАТЕГОРИЯ × ПРОДУКТ"
    write_matrix(ws, 47, 2, [["Категория", "Клиентов", *[row[0] for row in built["products"][:10]]], *built["heatmap"][:20]], style_row=48)
    ws.sheet_view.showGridLines = False

    # Категории
    ws = wb["Категории"]
    clear_sheet_values(ws, 1, 220, 18)
    ws["A1"] = "Категории"
    ws["A2"] = f"{len(built['categories'])} ниш. Где сильнее, где живём дольше, что внутри"
    cat_header = ["Категория", "Проектов", "Доля", "Подкатегорий", "Топ-подкатегория", "Топ-продукт", "Активных в 2025–26", "Ср. месяцев работы"]
    cat_rows = [cat_header, *built["categories"]]
    write_matrix(ws, 4, 1, cat_rows, style_row=5)
    for row in ws.iter_rows(min_row=5, max_row=4 + len(built["categories"]), min_col=3, max_col=3):
        row[0].number_format = "0.0%"
    for row in ws.iter_rows(min_row=5, max_row=4 + len(built["categories"]), min_col=8, max_col=8):
        row[0].number_format = "0.0"
    recreate_table(ws, "tbl_categories", f"A4:H{4 + len(built['categories'])}")
    style_header(ws, 4, 8)

    # Продукты
    ws = wb["Продукты"]
    clear_sheet_values(ws, 1, 220, 24)
    ws["A1"] = "Продукты"
    ws["A2"] = f"{len(built['products'])} продуктов. Счёт идёт по уникальным проектам, не по суммам оплат"
    prod_header = ["Продукт", "Проектов всего", "Доля", "Топ-категория", *[str(year) for year in years], "Δ 2024→2025"]
    prod_rows = [prod_header, *built["products"]]
    write_matrix(ws, 4, 1, prod_rows, style_row=5)
    for row in ws.iter_rows(min_row=5, max_row=4 + len(built["products"]), min_col=3, max_col=3):
        row[0].number_format = "0.0%"
    recreate_table(ws, "tbl_products", f"A4:{chr(64 + len(prod_header))}{4 + len(built['products'])}")
    style_header(ws, 4, len(prod_header))

    # Клиенты
    ws = wb["Клиенты"]
    clear_sheet_values(ws, 1, 900, 28)
    ws["A1"] = "Клиенты"
    ws["A2"] = f"{len(built['clients'])} проектов · поиск, фильтры, статус активности"
    ws["A4"] = "Поиск проекта →"
    ws["D4"] = "Услуги:"
    ws["C5"] = "Категория:"
    ws["E5"] = "Период:"
    ws["G5"] = "Месяцев:"
    ws["I5"] = "Поступлений:"
    ws["K5"] = "Статус:"
    client_header = ["Проект", "Категория", "Подкатегория", "Период", "Услуги", *[str(year) for year in years], "Месяцев", "Поступлений", "Статус"]
    client_rows = [client_header, *built["clients"]]
    write_matrix(ws, 7, 1, client_rows, style_row=8)
    last_client_row = 7 + len(built["clients"])
    recreate_table(ws, "tbl_clients", f"A7:{chr(64 + len(client_header))}{last_client_row}")
    style_header(ws, 7, len(client_header))
    ws.freeze_panes = "B8"

    # Служебные вкладки
    ws = wb["Продукт × Клиент"]
    clear_sheet_values(ws, 1, 2000, 18)
    pc_header = ["Продукт", "Проект", "Категория", "Подкатегория", "Годы продукта", "Период продукта", "Годы сотрудничества", "Все продукты проекта", "Месяцев", "Поступлений", "Статус данных"]
    write_matrix(ws, 1, 1, [pc_header, *built["product_client"]], style_row=2)
    recreate_table(ws, "tbl_product_client", f"A1:K{1 + len(built['product_client'])}")
    style_header(ws, 1, 11)

    ws = wb["Классификация"]
    clear_sheet_values(ws, 1, 1000, 12)
    cls_header = ["Проект", "Категория", "Подкатегория", "Источник", "Строк в исходнике", "Строки исходника"]
    write_matrix(ws, 1, 1, [cls_header, *built["classification"]], style_row=2)
    recreate_table(ws, "tbl_classification", f"A1:F{1 + len(built['classification'])}")
    style_header(ws, 1, 6)

    ws = wb["Данные"]
    clear_sheet_values(ws, 1, 10000, 8)
    db_rows = load_values(DB_JSON, "Данные!A:H")
    write_matrix(ws, 1, 1, db_rows, style_row=2)
    recreate_table(ws, "tbl_data", f"A1:H{len(db_rows)}")
    style_header(ws, 1, 8)

    ws = wb["dash_data"]
    clear_sheet_values(ws, 1, 300, 32)
    write_matrix(ws, 1, 1, [["Год", *built["top_products"], "Другие"], *built["dynamics"]])
    write_matrix(ws, 15, 1, [["Категория", "Клиентов", *[row[0] for row in built["products"][:10]]], *built["heatmap"]])
    write_matrix(ws, 45, 1, [["Когорта", "Размер", *years], *built["retention"]])

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
    rebuild_charts(wb)
    wb.save(OUTPUT)
    return OUTPUT


if __name__ == "__main__":
    print(update_workbook())
